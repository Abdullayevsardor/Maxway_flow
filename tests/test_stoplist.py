"""Стоп-лист (Причины стопов по продуктам по филиалам) — backend testlari.

Qamrov: yaratish, ro'yxat, detal, tahrirlash, filtr, saralash, pagination,
ruxsatlar (403) va Excel eksport.
"""
import io

import pytest

from conftest import login
from app import models


# ---------- yordamchilar ----------
def api_create(client, product_id, branch_id=None, reason="supplier_no_product",
               comment="Поставка задерживается", **kw):
    body = {"product_id": product_id, "reason": reason, "branch_comment": comment}
    if branch_id:
        body["branch_id"] = branch_id
    body.update(kw)
    return client.post("/api/product-stops", json=body)


# ================= SCENARIO 2 — создание =================
def test_create_sets_created_by_and_at_on_backend(client, seed, db):
    """Filial yozuv yaratadi; created_at/created_by faqat backend tomonidan qo'yiladi."""
    login(client, seed["branch"])
    r = api_create(client, seed["dishes"][0].id,
                   # frontend soxta qiymat yubormoqchi — backend e'tiborsiz qoldirishi kerak
                   created_by=seed["admin"].id, created_at="2000-01-01 00:00")
    assert r.status_code == 201, r.text
    item = r.json()["created"][0]
    assert item["branch"] == "Ресторан №12"
    assert item["product"] == "Бургер Классик"
    assert item["reason_label"] == "Нет продукта у поставщика"
    assert item["created_by"] == seed["branch"].id      # soxta created_by qabul qilinmadi
    assert not item["created_at"].startswith("2000")   # soxta created_at qabul qilinmadi
    assert item["supply_confirmed"] is False


def test_create_appears_in_list(client, seed):
    login(client, seed["branch"])
    r = client.get("/api/product-stops")
    assert r.status_code == 200
    names = [i["product"] for i in r.json()["items"]]
    assert "Бургер Классик" in names


def test_duplicate_is_skipped_not_duplicated(client, seed):
    """Bir filialda ayni taom ikki marta faol stopga tushmaydi."""
    login(client, seed["branch"])
    r = api_create(client, seed["dishes"][0].id)
    assert r.status_code == 201
    assert r.json()["created"] == []
    assert r.json()["skipped"] == ["Бургер Классик"]


# ================= SCENARIO 3 — validatsiya =================
@pytest.mark.parametrize("payload, expect", [
    ({"reason": "supplier_no_product"}, "Блюдо"),
    ({"product_id": 1}, "Причина"),
    ({"product_id": 1, "reason": "не_существует"}, "справочник"),
    ({"product_id": 999999, "reason": "supplier_no_product"}, "Блюдо не найдено"),
])
def test_validation_rejects_bad_input(client, seed, payload, expect):
    login(client, seed["branch"])
    r = client.post("/api/product-stops", json=payload)
    assert r.status_code == 400
    assert expect.lower() in r.json()["detail"].lower()


def test_validation_rejects_too_long_comment(client, seed):
    login(client, seed["branch"])
    r = api_create(client, seed["dishes"][1].id, comment="x" * 1001)
    assert r.status_code == 400
    assert "максимум" in r.json()["detail"]


def test_admin_must_pass_existing_branch(client, seed):
    login(client, seed["admin"])
    r = client.post("/api/product-stops", json={
        "product_id": seed["dishes"][1].id, "reason": "wrong_order", "branch_id": 999999})
    assert r.status_code == 400
    assert "Филиал не найден" in r.json()["detail"]


# ================= SCENARIO 4 — просмотр =================
def test_detail_shows_all_fields(client, seed, db):
    login(client, seed["branch"])
    sid = client.get("/api/product-stops").json()["items"][0]["id"]
    r = client.get(f"/api/product-stops/{sid}")
    assert r.status_code == 200
    d = r.json()
    for key in ("id", "created_at", "branch", "product", "reason_label",
                "branch_comment", "supply_confirmed", "supply_comment",
                "created_by_name", "updated_at"):
        assert key in d
    # HTML detal sahifasi ham ochiladi
    page = client.get(f"/stoplist/{sid}")
    assert page.status_code == 200
    assert "Информация о стопе" in page.text


def test_branch_cannot_view_other_branch_record(client, seed):
    """Filial boshqa filial yozuvini ko'ra olmaydi (403)."""
    login(client, seed["admin"])
    r = api_create(client, seed["dishes"][2].id, branch_id=seed["b2"].id,
                   reason="equipment_broken")
    other_id = r.json()["created"][0]["id"]
    login(client, seed["branch"])                    # 12-filial
    assert client.get(f"/api/product-stops/{other_id}").status_code == 403
    assert client.get(f"/stoplist/{other_id}").status_code == 403


def test_branch_list_is_scoped_to_own_branch(client, seed):
    login(client, seed["branch"])
    items = client.get("/api/product-stops").json()["items"]
    assert items and all(i["branch"] == "Ресторан №12" for i in items)


# ================= SCENARIO 5 — редактирование =================
def test_edit_updates_updated_at_and_by(client, seed):
    login(client, seed["branch"])
    sid = client.get("/api/product-stops").json()["items"][0]["id"]
    assert client.get(f"/api/product-stops/{sid}").json()["updated_at"] is None
    r = client.patch(f"/api/product-stops/{sid}",
                     json={"branch_comment": "Поставка сегодня к 16:00"})
    assert r.status_code == 200, r.text
    d = r.json()
    assert d["branch_comment"] == "Поставка сегодня к 16:00"
    assert d["updated_at"] is not None
    assert d["updated_by"] == seed["branch"].id


def test_edit_rejects_unknown_reason(client, seed):
    login(client, seed["branch"])
    sid = client.get("/api/product-stops").json()["items"][0]["id"]
    r = client.patch(f"/api/product-stops/{sid}", json={"reason": "нет_такой"})
    assert r.status_code == 400


# ================= SCENARIO 6 + 11 — подтверждение снабжения =================
def test_supply_can_confirm(client, seed):
    login(client, seed["supply"])
    sid = client.get("/api/product-stops").json()["items"][0]["id"]
    r = client.patch(f"/api/product-stops/{sid}", json={
        "supply_confirmed": True, "supply_comment": "Поставка ожидается сегодня до 16:00"})
    assert r.status_code == 200, r.text
    d = r.json()
    assert d["supply_confirmed"] is True
    assert d["supply_comment"] == "Поставка ожидается сегодня до 16:00"
    assert d["confirmed_at"] is not None
    assert d["confirmed_by_name"] == "supply"


def test_branch_cannot_confirm_via_direct_api(client, seed):
    """SCENARIO 11 — ruxsatsiz foydalanuvchi to'g'ridan-to'g'ri API orqali 403 oladi."""
    login(client, seed["branch"])
    sid = client.get("/api/product-stops").json()["items"][0]["id"]
    before = client.get(f"/api/product-stops/{sid}").json()["supply_confirmed"]
    r = client.patch(f"/api/product-stops/{sid}", json={"supply_confirmed": True})
    assert r.status_code == 403
    # forma endpointi ham himoyalangan
    r2 = client.post(f"/stoplist/{sid}/confirm", data={"supply_confirmed": "1"})
    assert r2.status_code == 403
    # va qiymat o'zgarmagan
    assert client.get(f"/api/product-stops/{sid}").json()["supply_confirmed"] is before


def test_supply_cannot_edit_branch_fields(client, seed):
    login(client, seed["supply"])
    sid = client.get("/api/product-stops").json()["items"][0]["id"]
    r = client.patch(f"/api/product-stops/{sid}", json={"branch_comment": "чужое поле"})
    assert r.status_code == 403


def test_viewer_cannot_create_or_edit(client, seed):
    login(client, seed["viewer"])
    assert api_create(client, seed["dishes"][3].id).status_code == 403
    sid = client.get("/api/product-stops").json()["items"][0]["id"]
    assert client.patch(f"/api/product-stops/{sid}",
                        json={"supply_comment": "x"}).status_code == 403


def test_client_without_branch_sees_nothing(client, seed, db):
    """Filialga biriktirilmagan klient hech qanday yozuv ko'rmaydi."""
    from app import auth as _auth
    u = models.User(full_name="no-branch", email="nobranch@t.uz",
                    hashed_password=_auth.hash_password("test12345"),
                    role=models.Role.client, is_active=True)
    db.add(u); db.commit()
    login(client, u)
    assert client.get("/api/product-stops").json()["total"] == 0
    assert client.get("/stoplist").status_code == 200


def test_anonymous_gets_401(client, seed):
    client.cookies.clear()
    assert client.get("/api/product-stops").status_code == 401
    assert client.post("/api/product-stops", json={}).status_code == 401


# ================= SCENARIO 7 — фильтрация =================
def test_filter_by_branch_reason_and_confirmation(client, seed):
    login(client, seed["admin"])
    # 7-filialga yana bitta yozuv
    api_create(client, seed["dishes"][3].id, branch_id=seed["b2"].id, reason="menu_removed")

    by_branch = client.get(f"/api/product-stops?branch_id={seed['b2'].id}").json()
    assert by_branch["items"] and all(i["branch"] == "Ресторан №7" for i in by_branch["items"])

    by_reason = client.get("/api/product-stops?reason=menu_removed").json()
    assert all(i["reason"] == "menu_removed" for i in by_reason["items"])
    assert by_reason["total"] >= 1

    conf_yes = client.get("/api/product-stops?confirmed=yes").json()
    assert all(i["supply_confirmed"] is True for i in conf_yes["items"])
    conf_no = client.get("/api/product-stops?confirmed=no").json()
    assert all(i["supply_confirmed"] is False for i in conf_no["items"])
    assert conf_yes["total"] + conf_no["total"] == client.get("/api/product-stops").json()["total"]


def test_filter_by_product_and_date_range(client, seed):
    login(client, seed["admin"])
    pid = seed["dishes"][0].id
    d = client.get(f"/api/product-stops?menu_item_id={pid}").json()
    assert d["items"] and all(i["product_id"] == pid for i in d["items"])

    today = models.tashkent_now().strftime("%Y-%m-%d")
    assert client.get(f"/api/product-stops?date_from={today}&date_to={today}").json()["total"] >= 1
    assert client.get("/api/product-stops?date_from=1999-01-01&date_to=1999-01-02").json()["total"] == 0


def test_invalid_filter_is_ignored_not_crashing(client, seed):
    login(client, seed["admin"])
    r = client.get("/api/product-stops?branch_id=abc&reason=' OR 1=1--&date_from=не-дата")
    assert r.status_code == 200
    assert r.json()["total"] == client.get("/api/product-stops").json()["total"]


# ================= SCENARIO 8 — сортировка =================
def test_sorting_asc_and_desc(client, seed):
    login(client, seed["admin"])
    asc = [i["product"] for i in
           client.get("/api/product-stops?sort_by=dish&sort_order=asc").json()["items"]]
    desc = [i["product"] for i in
            client.get("/api/product-stops?sort_by=dish&sort_order=desc").json()["items"]]
    assert asc == sorted(asc)
    assert desc == list(reversed(asc))

    dates = [i["created_at"] for i in
             client.get("/api/product-stops?sort_by=created_at&sort_order=asc").json()["items"]]
    assert dates == sorted(dates)


def test_unknown_sort_field_falls_back(client, seed):
    """Oq ro'yxatda yo'q ustun — standart tartibga qaytadi (SQL injection imkonsiz)."""
    login(client, seed["admin"])
    r = client.get("/api/product-stops?sort_by=id);DROP TABLE stop_entries;--")
    assert r.status_code == 200
    assert r.json()["sort_by"] == "created_at"


# ================= SCENARIO 9 — pagination =================
def test_pagination_meta_and_slicing(client, seed):
    login(client, seed["admin"])
    total = client.get("/api/product-stops").json()["total"]
    assert total >= 3
    p1 = client.get("/api/product-stops?page=1&page_size=10&sort_by=created_at&sort_order=asc").json()
    assert p1["page"] == 1 and p1["page_size"] == 10
    assert p1["pages"] == max(1, -(-total // 10))
    assert len(p1["items"]) <= 10

    small = client.get("/api/product-stops?page=1&page_size=10").json()
    assert small["page_size"] == 10
    # ruxsat etilmagan page_size — standartga qaytadi
    assert client.get("/api/product-stops?page_size=7").json()["page_size"] == 25
    # chegaradan tashqari sahifa — oxirgi sahifaga qisiladi
    assert client.get("/api/product-stops?page=9999").json()["page"] == small["pages"]
    assert client.get("/api/product-stops?page=-3").json()["page"] == 1


def test_pages_do_not_overlap(client, seed):
    login(client, seed["admin"])
    base = "/api/product-stops?page_size=10&sort_by=created_at&sort_order=asc&page="
    ids1 = [i["id"] for i in client.get(base + "1").json()["items"]]
    meta = client.get(base + "1").json()
    if meta["pages"] > 1:
        ids2 = [i["id"] for i in client.get(base + "2").json()["items"]]
        assert not set(ids1) & set(ids2)


# ================= SCENARIO 10 — Excel =================
def _read_xlsx(content):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb["Все филиалы"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return rows


def test_export_respects_filters(client, seed):
    login(client, seed["admin"])
    r = client.get(f"/stoplist/export?mode=active&branch_id={seed['b2'].id}")
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    assert "product_stops_" in r.headers["content-disposition"]

    rows = _read_xlsx(r.content)
    assert rows[0] == ("Добавлено", "Филиал", "Блюдо", "Причина", "Комментарий Филиала",
                       "Подтверждение причины стопа отделом снабжения", "Комментарий Снабжения")
    body = rows[1:]
    assert body, "экспорт не должен быть пустым"
    assert all(row[1] == "Ресторан №7" for row in body)   # faqat filtrlangan filial
    assert all(row[5] in ("ДА", "НЕТ") for row in body)


def test_export_respects_sorting(client, seed):
    login(client, seed["admin"])
    rows = _read_xlsx(client.get(
        "/stoplist/export?mode=active&sort_by=dish&sort_order=asc").content)[1:]
    dishes = [r[2] for r in rows]
    assert dishes == sorted(dishes)


def test_export_forbidden_without_permission(client, seed, db):
    """export_stop ruxsati olib tashlangan foydalanuvchi 403 oladi."""
    import json as _json
    u = seed["viewer"]
    u.perms = _json.dumps({"export_stop": False})
    db.commit()
    login(client, u)
    assert client.get("/stoplist/export?mode=active").status_code == 403
    u.perms = None
    db.commit()


# ================= HTML sahifalar =================
def test_html_pages_render(client, seed):
    login(client, seed["admin"])
    for url in ("/stoplist", "/stoplist/history",
                "/stoplist?sort_by=dish&sort_order=asc&page_size=10",
                "/stoplist?confirmed=yes&reason=menu_removed"):
        r = client.get(url)
        assert r.status_code == 200, url
        assert "Причины стопов" in r.text or "История стоп-листа" in r.text


def test_html_add_form_creates_record(client, seed):
    """Forma orqali qo'shish (klient) — ro'yxatда paydo bo'ladi."""
    login(client, seed["branch"])
    r = client.post("/stoplist/add", data={
        "menu_item_id": [str(seed["dishes"][2].id)],
        "reason": "equipment_broken", "comment": "Сломался аппарат"},
        follow_redirects=False)
    assert r.status_code == 302
    assert "ok=" in r.headers["location"]
    items = client.get("/api/product-stops").json()["items"]
    assert any(i["product"] == "Кола 0.5" and i["reason"] == "equipment_broken"
               for i in items)


def test_html_add_without_reason_is_rejected(client, seed):
    login(client, seed["branch"])
    r = client.post("/stoplist/add", data={
        "menu_item_id": [str(seed["dishes"][3].id)], "reason": ""},
        follow_redirects=False)
    assert r.status_code == 302
    assert "err=" in r.headers["location"]


def test_html_edit_form(client, seed):
    login(client, seed["supply"])
    sid = client.get("/api/product-stops").json()["items"][0]["id"]
    r = client.post(f"/stoplist/{sid}/edit", data={
        "fields": "supply_confirmed,supply_comment",
        "supply_confirmed": "1", "supply_comment": "Ждём поставку"},
        follow_redirects=False)
    assert r.status_code == 302
    d = client.get(f"/api/product-stops/{sid}").json()
    assert d["supply_confirmed"] is True and d["supply_comment"] == "Ждём поставку"


def test_resolve_moves_to_history(client, seed):
    login(client, seed["branch"])
    before = client.get("/api/product-stops").json()["items"]
    sid = before[0]["id"]
    r = client.post(f"/stoplist/{sid}/resolve", follow_redirects=False)
    assert r.status_code == 302
    assert sid not in [i["id"] for i in client.get("/api/product-stops").json()["items"]]
    hist = client.get("/api/product-stops?mode=history").json()["items"]
    assert sid in [i["id"] for i in hist]


# ================= спрвочники =================
def test_reference_endpoints(client, seed):
    login(client, seed["admin"])
    reasons = client.get("/api/stop-reasons").json()
    keys = {r["id"] for r in reasons}
    assert {"equipment_broken", "supplier_no_product",
            "wrong_order", "menu_removed"} <= keys
    branch_names = {b["name"] for b in client.get("/api/branches").json()}
    assert {"Ресторан №12", "Ресторан №7"} <= branch_names
    dish_names = {p["name"] for p in client.get("/api/products").json()}
    assert {"Бургер Классик", "Кола 0.5"} <= dish_names

    # klient — faqat o'z filialini ko'radi
    login(client, seed["branch"])
    assert [b["name"] for b in client.get("/api/branches").json()] == ["Ресторан №12"]


# ================= PERFORMANCE — N+1 yo'qligi =================
def test_list_query_count_is_constant(client, seed, db):
    """Ro'yxat sahifasi yozuvlar soniga bog'liq bo'lmagan sonda so'rov qiladi."""
    from sqlalchemy import event
    from app.database import engine

    counter = {"n": 0}

    def before(conn, cursor, statement, params, ctx, many):
        counter["n"] += 1

    login(client, seed["admin"])
    event.listen(engine, "before_cursor_execute", before)
    try:
        counter["n"] = 0
        client.get("/api/product-stops?page_size=100")
        used = counter["n"]
    finally:
        event.remove(engine, "before_cursor_execute", before)
    # 1 — foydalanuvchi, 1 — count, 1 — ro'yxat (branch/menu_item join bilan keladi)
    assert used <= 5, f"слишком много запросов: {used} (возможен N+1)"


# ================= TELEGRAM XABARNOMASI =================
@pytest.fixture
def fresh_dish(db):
    """Har chaqiruvda hali stopga tushmagan yangi taom qaytaradi."""
    counter = {"n": 0}

    def _make():
        counter["n"] += 1
        m = models.MenuItem(name=f"Тест-блюдо {counter['n']}-{id(counter)}", is_active=True)
        db.add(m)
        db.commit()
        return m.id
    return _make


@pytest.fixture
def tg(monkeypatch):
    """send_telegram ni ushlaydi: (chat_id, matn) juftliklari to'planadi."""
    import main
    sent = []
    monkeypatch.setattr(main, "send_telegram",
                        lambda cid, text, **kw: sent.append((cid, text)))
    # fon oqimi o'rniga darhol yuboramiz — test barqaror bo'lsin
    monkeypatch.setattr(main, "_send_async",
                        lambda ids, text, button_url="", token="":
                        [main.send_telegram(c, text) for c in ids])
    return sent


def test_stop_notification_reaches_right_people(client, seed, db, tg, monkeypatch, fresh_dish):
    """Снабжение/просмотр/админ — barcha filiallar; filial — faqat o'zi."""
    import main
    b1, b2 = seed["b1"], seed["b2"]
    # har bir rolga telegram ID beramiz
    seed["supply"].telegram_chat_id = "SUPPLY"
    seed["viewer"].telegram_chat_id = "VIEWER"
    seed["admin"].telegram_chat_id = "ADMIN"
    seed["branch"].telegram_chat_id = "BRANCH12"      # b1 filiali
    seed["branch2"].telegram_chat_id = "BRANCH7"      # b2 filiali
    b1.tg_chat_ids = "EMP1, EMP2"                     # b1 xodimlari
    b2.tg_chat_ids = "EMP7"
    db.commit()
    monkeypatch.setattr(main, "get_stop_channel", lambda: "CHANNEL")

    # 12-filial o'z stopini qo'shadi
    login(client, seed["branch"])
    tg.clear()
    r = api_create(client, fresh_dish(), reason="menu_removed",
                   comment="Выводим из меню")
    assert r.status_code == 201, r.text
    got = {cid for cid, _ in tg}
    # barcha filiallarni kuzatuvchilar
    assert {"SUPPLY", "VIEWER", "ADMIN", "CHANNEL"} <= got
    # o'z filiali xodimlari
    assert {"EMP1", "EMP2"} <= got
    # boshqa filialga tegishlilar — YO'Q
    assert "BRANCH7" not in got and "EMP7" not in got
    # qo'shgan odamning o'ziga ham yuborilmaydi
    assert "BRANCH12" not in got

    text = tg[0][1]
    assert "Ресторан №12" in text
    assert "Вывод из меню продукта" in text
    assert "Выводим из меню" in text


def test_kpp_gets_only_assigned_branches(client, seed, db, tg, monkeypatch, fresh_dish):
    import main
    from app import auth as _auth
    monkeypatch.setattr(main, "get_stop_channel", lambda: "")
    kpp = models.User(full_name="kpp", email="kpp@t.uz",
                      hashed_password=_auth.hash_password("test12345"),
                      role=models.Role.kpp, is_active=True,
                      telegram_chat_id="KPP")
    kpp.visible_branches = [seed["b2"]]        # faqat 7-filial
    db.add(kpp); db.commit()

    login(client, seed["branch"])              # 12-filial qo'shadi
    tg.clear()
    api_create(client, fresh_dish(), reason="wrong_order")
    assert "KPP" not in {c for c, _ in tg}, "KPP begona filial xabarini oldi"

    login(client, seed["branch2"])             # 7-filial qo'shadi
    tg.clear()
    api_create(client, fresh_dish(), reason="wrong_order")
    assert "KPP" in {c for c, _ in tg}


def test_kpp_without_branches_gets_all(client, seed, db, tg, monkeypatch, fresh_dish):
    """Filial biriktirilmagan КПП — barcha filiallarni oladi."""
    import main
    from app import auth as _auth
    monkeypatch.setattr(main, "get_stop_channel", lambda: "")
    kpp = models.User(full_name="kpp2", email="kpp2@t.uz",
                      hashed_password=_auth.hash_password("test12345"),
                      role=models.Role.kpp, is_active=True, telegram_chat_id="KPPALL")
    db.add(kpp); db.commit()
    login(client, seed["branch"])
    tg.clear()
    api_create(client, fresh_dish(), reason="equipment_broken")
    assert "KPPALL" in {c for c, _ in tg}


def test_one_message_for_many_dishes(client, seed, db, tg, monkeypatch, fresh_dish):
    """Bir nechta taom birga qo'shilsa — bitta umumiy xabar."""
    import main
    monkeypatch.setattr(main, "get_stop_channel", lambda: "")
    seed["supply"].telegram_chat_id = "SUPPLY"
    db.commit()
    login(client, seed["branch2"])
    tg.clear()
    ids = [fresh_dish(), fresh_dish()]
    r = client.post("/api/product-stops", json={
        "product_ids": ids, "reason": "supplier_no_product",
        "branch_comment": "Поставка задерживается"})
    assert r.status_code == 201, r.text
    to_supply = [t for c, t in tg if c == "SUPPLY"]
    assert len(to_supply) == 1, "har bir taom uchun alohida xabar ketmasligi kerak"
    assert "Блюда (2)" in to_supply[0]


def test_inactive_and_no_chat_id_are_skipped(client, seed, db, tg, monkeypatch, fresh_dish):
    import main
    monkeypatch.setattr(main, "get_stop_channel", lambda: "")
    seed["viewer"].telegram_chat_id = "VIEWER"
    seed["viewer"].is_active = False           # o'chirilgan foydalanuvchi
    seed["supply"].telegram_chat_id = ""       # ID kiritilmagan
    db.commit()
    login(client, seed["branch"])
    tg.clear()
    api_create(client, fresh_dish(), reason="equipment_broken")
    got = {c for c, _ in tg}
    assert "VIEWER" not in got and "" not in got
    seed["viewer"].is_active = True
    db.commit()


def test_notification_on_resolve(client, seed, db, tg, monkeypatch, fresh_dish):
    """Stopdan olinganda ham xabar ketadi — qo'shilgandagi manzillarga."""
    import main
    monkeypatch.setattr(main, "get_stop_channel", lambda: "CHANNEL")
    seed["supply"].telegram_chat_id = "SUPPLY"
    seed["viewer"].telegram_chat_id = "VIEWER"
    seed["branch2"].telegram_chat_id = "BRANCH7"
    db.commit()

    login(client, seed["branch"])
    r = api_create(client, fresh_dish(), reason="equipment_broken")
    sid = r.json()["created"][0]["id"]

    tg.clear()
    assert client.post(f"/stoplist/{sid}/resolve",
                       follow_redirects=False).status_code == 302
    got = {c for c, _ in tg}
    assert {"SUPPLY", "VIEWER", "CHANNEL"} <= got
    assert "BRANCH7" not in got          # boshqa filialga tegishli emas

    text = [t for c, t in tg if c == "SUPPLY"][0]
    assert "Снят со стопа" in text
    assert "Ресторан №12" in text
    assert "Сломалось оборудование" in text
    assert "На стопе был" in text


def test_no_notification_when_resolve_forbidden(client, seed, db, tg, monkeypatch, fresh_dish):
    """Ruxsatsiz urinishda 403 va hech qanday xabar ketmaydi."""
    import main
    monkeypatch.setattr(main, "get_stop_channel", lambda: "CHANNEL")
    login(client, seed["admin"])
    r = api_create(client, fresh_dish(), branch_id=seed["b2"].id, reason="wrong_order")
    sid = r.json()["created"][0]["id"]

    login(client, seed["viewer"])         # viewer stopdan ololmaydi
    tg.clear()
    assert client.post(f"/stoplist/{sid}/resolve").status_code == 403
    assert tg == []


@pytest.mark.parametrize("mins, expect", [
    (0, "меньше минуты"),
    (12, "12 мин"),
    (220, "3 ч 40 мин"),
    (60 * 24 * 2 + 60 * 5, "2 д 5 ч"),
])
def test_duration_wording(mins, expect):
    import main
    from datetime import timedelta
    assert main._human_duration(timedelta(minutes=mins)) == expect


def test_stop_notifications_use_separate_bot_token(monkeypatch, client, seed, db, fresh_dish):
    """Stop xabarlari MAXWAY_STOP_BOT_TOKEN bilan ketadi, boshqalari — asosiy token bilan."""
    import main
    used = []
    monkeypatch.setattr(main, "send_telegram",
                        lambda cid, text, button_url="", button_text="", token="":
                        used.append((cid, token)))
    monkeypatch.setattr(main, "_send_async",
                        lambda ids, text, button_url="", token="":
                        [main.send_telegram(c, text, token=token) for c in ids])
    monkeypatch.setattr(main, "get_bot_token", lambda: "MAIN-TOKEN")
    monkeypatch.setattr(main, "get_stop_channel", lambda: "CHANNEL")
    monkeypatch.setenv("MAXWAY_STOP_BOT_TOKEN", "STOP-TOKEN")

    login(client, seed["branch"])
    api_create(client, fresh_dish(), reason="wrong_order")
    assert used, "xabar yuborilmadi"
    assert all(tok == "STOP-TOKEN" for _, tok in used), used


def test_stop_token_falls_back_to_main(monkeypatch):
    """Alohida token sozlanmagan bo'lsa — asosiy bot ishlatiladi."""
    import main
    monkeypatch.delenv("MAXWAY_STOP_BOT_TOKEN", raising=False)
    monkeypatch.setattr(main, "get_bot_token", lambda: "MAIN-TOKEN")
    monkeypatch.chdir(main.os.path.dirname(main.os.path.abspath(main.__file__)))
    assert main.get_stop_bot_token() == "MAIN-TOKEN"


def test_notify_preview_is_admin_only(client, seed, db):
    """Tashxis endpointi faqat admin uchun va hech nima yubormaydi."""
    login(client, seed["branch"])
    assert client.get(f"/api/stop-notify-preview?branch_id={seed['b1'].id}").status_code == 403
    login(client, seed["supply"])
    assert client.get(f"/api/stop-notify-preview?branch_id={seed['b1'].id}").status_code == 403
    login(client, seed["admin"])
    assert client.get("/api/stop-notify-preview?branch_id=999999").status_code == 404


def test_notify_preview_shows_sources(client, seed, db, monkeypatch):
    import main
    monkeypatch.setattr(main, "get_stop_channel", lambda: "CHANNEL")
    seed["supply"].telegram_chat_id = "SUPPLY"
    seed["branch"].telegram_chat_id = "BRANCH12"
    seed["b1"].tg_chat_ids = "EMP1, EMP2"
    db.commit()
    login(client, seed["admin"])
    r = client.get(f"/api/stop-notify-preview?branch_id={seed['b1'].id}&check=0")
    assert r.status_code == 200, r.text
    d = r.json()
    assert d["branch"] == "Ресторан №12"
    got = {x["chat_id"]: x["source"] for x in d["recipients"]}
    assert got["SUPPLY"] == "Снабжение"
    assert got["BRANCH12"] == "Логин филиала"
    assert got["EMP1"] == "Сотрудник филиала"
    assert got["EMP2"] == "Сотрудник филиала"
    assert got["CHANNEL"] == "Канал"


# ================= OMMAVIY OLIB TASHLASH =================
def test_bulk_resolve_removes_all_selected(client, seed, db, tg, monkeypatch, fresh_dish):
    """10 ta taomni bir vaqtda stopdan olish — bitta umumiy telegram xabari."""
    import main
    monkeypatch.setattr(main, "get_stop_channel", lambda: "")
    seed["supply"].telegram_chat_id = "SUPPLY"
    db.commit()

    login(client, seed["branch"])
    ids = [fresh_dish() for _ in range(10)]
    r = client.post("/api/product-stops", json={
        "product_ids": ids, "reason": "supplier_no_product",
        "branch_comment": "Нет сыра"})
    assert r.status_code == 201, r.text
    sids = [x["id"] for x in r.json()["created"]]
    assert len(sids) == 10

    tg.clear()
    resp = client.post("/stoplist/resolve-bulk",
                       data={"sid": [str(i) for i in sids]},
                       follow_redirects=False)
    assert resp.status_code == 302
    assert "ok=" in resp.headers["location"]

    # hammasi tarixga o'tdi
    active = {i["id"] for i in client.get("/api/product-stops").json()["items"]}
    assert not (set(sids) & active)
    hist = {i["id"] for i in client.get("/api/product-stops?mode=history&page_size=100").json()["items"]}
    assert set(sids) <= hist

    # bitta umumiy xabar, 10 ta emas
    to_supply = [t for c, t in tg if c == "SUPPLY"]
    assert len(to_supply) == 1, f"{len(to_supply)} ta xabar ketdi"
    assert "Блюда (10)" in to_supply[0]
    assert "Снят со стопа" in to_supply[0]


def test_bulk_resolve_checks_permission_per_entry(client, seed, db, fresh_dish):
    """Begona filial yozuvi ro'yxatga qo'shilsa — 403, hech nima o'zgarmaydi."""
    login(client, seed["admin"])
    mine = client.post("/api/product-stops", json={
        "product_ids": [fresh_dish()], "reason": "wrong_order",
        "branch_id": seed["b1"].id}).json()["created"][0]["id"]
    other = client.post("/api/product-stops", json={
        "product_ids": [fresh_dish()], "reason": "wrong_order",
        "branch_id": seed["b2"].id}).json()["created"][0]["id"]

    login(client, seed["branch"])          # 12-filial
    r = client.post("/stoplist/resolve-bulk",
                    data={"sid": [str(mine), str(other)]})
    assert r.status_code == 403
    # ikkalasi ham stopda qolgan
    assert client.get(f"/api/product-stops/{mine}").json()["resolved"] is False


def test_bulk_resolve_empty_selection(client, seed):
    login(client, seed["branch"])
    r = client.post("/stoplist/resolve-bulk", data={}, follow_redirects=False)
    assert r.status_code == 302
    assert "err=" in r.headers["location"]


def test_bulk_resolve_groups_message_per_branch(client, seed, db, tg, monkeypatch, fresh_dish):
    """Admin turli filialdan olsa — har bir filialga o'z xabari."""
    import main
    monkeypatch.setattr(main, "get_stop_channel", lambda: "")
    seed["viewer"].telegram_chat_id = "VIEWER"
    db.commit()
    login(client, seed["admin"])
    a = client.post("/api/product-stops", json={
        "product_ids": [fresh_dish()], "reason": "menu_removed",
        "branch_id": seed["b1"].id}).json()["created"][0]["id"]
    b = client.post("/api/product-stops", json={
        "product_ids": [fresh_dish()], "reason": "menu_removed",
        "branch_id": seed["b2"].id}).json()["created"][0]["id"]
    tg.clear()
    client.post("/stoplist/resolve-bulk", data={"sid": [str(a), str(b)]})
    texts = [t for c, t in tg if c == "VIEWER"]
    assert len(texts) == 2, "har bir filial uchun alohida xabar kutilgan"
    assert any("Ресторан №12" in t for t in texts)
    assert any("Ресторан №7" in t for t in texts)


def test_branch_cannot_resolve_other_branch_single(client, seed, db, fresh_dish):
    """Filial BOSHQA filial yozuvini yakka tartibda ham stopdan ololmaydi."""
    login(client, seed["admin"])
    other = client.post("/api/product-stops", json={
        "product_ids": [fresh_dish()], "reason": "wrong_order",
        "branch_id": seed["b2"].id}).json()["created"][0]["id"]
    login(client, seed["branch"])          # 12-filial
    assert client.post(f"/stoplist/{other}/resolve").status_code == 403
    login(client, seed["admin"])
    assert client.get(f"/api/product-stops/{other}").json()["resolved"] is False


# ================= FILIALNI O'CHIRISH =================
def test_branch_delete_cleans_up_everything(client, seed, db, fresh_dish, monkeypatch):
    """Filial o'chirilganda: zayavka saqlanadi, stop o'chadi, login uziladi."""
    import main
    from app import auth as _auth
    monkeypatch.setattr(main, "get_stop_channel", lambda: "")

    b = models.Branch(name="Удаляемый филиал")
    db.add(b); db.flush()
    cli = models.User(full_name="del-login", email="del-login@t.uz",
                      hashed_password=_auth.hash_password("test12345"),
                      role=models.Role.client, user_branch_id=b.id, is_active=True)
    db.add(cli); db.flush()
    dep = db.query(models.Department).first()
    req = models.Request(title="Заявка филиала", department_id=dep.id,
                         created_by=cli.id, branch_id=b.id)
    db.add(req); db.flush()
    kpp = models.User(full_name="kpp-del", email="kpp-del@t.uz",
                      hashed_password=_auth.hash_password("test12345"),
                      role=models.Role.kpp, is_active=True)
    kpp.visible_branches = [b]
    db.add(kpp); db.commit()
    bid, rid, cid = b.id, req.id, cli.id

    login(client, seed["admin"])
    client.post("/api/product-stops", json={
        "product_ids": [fresh_dish(), fresh_dish()], "reason": "wrong_order",
        "branch_id": bid})
    assert db.query(models.StopEntry).filter(models.StopEntry.branch_id == bid).count() == 2

    r = client.post(f"/admin/branches/{bid}/delete", follow_redirects=False)
    assert r.status_code == 302, r.text
    db.expire_all()

    # filial o'chdi
    assert db.get(models.Branch, bid) is None
    # stop yozuvlari o'chdi — yetim qolmadi
    assert db.query(models.StopEntry).filter(models.StopEntry.branch_id == bid).count() == 0
    # zayavkani filial logini yaratgan edi — login bilan birga o'chadi
    assert db.get(models.Request, rid) is None
    # filial logini ham o'chdi (filial va login bir butun)
    assert db.get(models.User, cid) is None
    # КПП bog'lanishi tozalandi
    db.refresh(kpp)
    assert [x.id for x in kpp.visible_branches] == []


def test_branch_delete_is_admin_only(client, seed, db):
    b = models.Branch(name="Чужой филиал")
    db.add(b); db.commit()
    login(client, seed["branch"])
    assert client.post(f"/admin/branches/{b.id}/delete",
                       follow_redirects=False).status_code == 302
    assert db.get(models.Branch, b.id) is not None      # o'chmagan


def test_delete_user_also_deletes_its_branch(client, seed, db, fresh_dish):
    """Пользователи qismidan filial loginini o'chirsa — filial ham o'chadi (500 emas)."""
    from app import auth as _auth
    b = models.Branch(name="Парный филиал")
    db.add(b); db.flush()
    cli = models.User(full_name="pair-login", email="pair@t.uz",
                      hashed_password=_auth.hash_password("test12345"),
                      role=models.Role.client, user_branch_id=b.id, is_active=True)
    db.add(cli); db.commit()
    bid, uid = b.id, cli.id

    # filial loginidan stop qo'shamiz — aynan shu 500 xatoga sabab bo'lardi
    login(client, cli)
    r = client.post("/api/product-stops", json={
        "product_ids": [fresh_dish()], "reason": "wrong_order"})
    assert r.status_code == 201, r.text

    login(client, seed["admin"])
    resp = client.post(f"/admin/users/{uid}/delete", follow_redirects=False)
    assert resp.status_code == 302, f"500 qaytdi: {resp.status_code}"
    db.expire_all()
    assert db.get(models.User, uid) is None          # login o'chdi
    assert db.get(models.Branch, bid) is None        # filial ham o'chdi
    assert db.query(models.StopEntry).filter(
        models.StopEntry.branch_id == bid).count() == 0


def test_delete_ordinary_user_keeps_branches(client, seed, db):
    """Oddiy xodimni o'chirish filiallarга tegmaydi."""
    from app import auth as _auth
    u = models.User(full_name="обычный", email="ordinary@t.uz",
                    hashed_password=_auth.hash_password("test12345"),
                    role=models.Role.executor, is_active=True)
    db.add(u); db.commit()
    uid = u.id
    before = db.query(models.Branch).count()
    login(client, seed["admin"])
    assert client.post(f"/admin/users/{uid}/delete",
                       follow_redirects=False).status_code == 302
    db.expire_all()
    assert db.get(models.User, uid) is None
    assert db.query(models.Branch).count() == before


def test_admin_cannot_delete_self(client, seed, db):
    login(client, seed["admin"])
    r = client.post(f"/admin/users/{seed['admin'].id}/delete", follow_redirects=False)
    assert r.status_code == 302 and "err=" in r.headers["location"]
    db.expire_all()
    assert db.get(models.User, seed["admin"].id) is not None


# ================= HIMOYALANGAN AKKAUNT =================
def test_protected_admin_cannot_be_deleted(client, seed, db, monkeypatch):
    """Bosh adminni boshqa admin ham o'chira olmaydi."""
    import main
    from app import auth as _auth
    boss = models.User(full_name="Главный", email="a.ruzikulov@maxway.uz",
                       hashed_password=_auth.hash_password("test12345"),
                       role=models.Role.admin, is_active=True)
    db.add(boss); db.commit()
    bid = boss.id
    assert main.is_protected_user(boss) is True

    login(client, seed["admin"])          # boshqa admin
    r = client.post(f"/admin/users/{bid}/delete", follow_redirects=False)
    assert r.status_code == 302
    assert "err=" in r.headers["location"]
    db.expire_all()
    assert db.get(models.User, bid) is not None, "himoyalangan akkaunt o'chib ketdi"


def test_protected_list_is_configurable(monkeypatch):
    import main
    monkeypatch.setenv("MAXWAY_PROTECTED_EMAILS", "boss@x.uz, second@x.uz")

    class U:
        def __init__(self, e): self.email = e
    assert main.is_protected_user(U("boss@x.uz")) is True
    assert main.is_protected_user(U("SECOND@X.UZ")) is True       # registr muhim emas
    assert main.is_protected_user(U("a.ruzikulov@maxway.uz")) is False
    assert main.is_protected_user(None) is False


def test_protected_user_survives_branch_delete(client, seed, db, monkeypatch):
    """Himoyalangan akkaunt filialga biriktirilgan bo'lsa ham o'chmaydi."""
    import main
    from app import auth as _auth
    monkeypatch.setenv("MAXWAY_PROTECTED_EMAILS", "keepme@x.uz")
    b = models.Branch(name="Филиал с боссом"); db.add(b); db.flush()
    boss = models.User(full_name="boss", email="keepme@x.uz",
                       hashed_password=_auth.hash_password("test12345"),
                       role=models.Role.client, user_branch_id=b.id, is_active=True)
    db.add(boss); db.commit()
    bid, uid = b.id, boss.id

    login(client, seed["admin"])
    assert client.post(f"/admin/branches/{bid}/delete",
                       follow_redirects=False).status_code == 302
    db.expire_all()
    assert db.get(models.Branch, bid) is None        # filial o'chdi
    u = db.get(models.User, uid)
    assert u is not None, "himoyalangan akkaunt filial bilan birga o'chib ketdi"
    assert u.user_branch_id is None                  # faqat uzildi


def test_any_branch_linked_user_deletes_branch(client, seed, db):
    """Roli client bo'lmasa ham, filialga biriktirilgan bo'lsa — filial ham o'chadi."""
    from app import auth as _auth
    b = models.Branch(name="Филиал менеджера"); db.add(b); db.flush()
    u = models.User(full_name="mgr", email="mgr-branch@t.uz",
                    hashed_password=_auth.hash_password("test12345"),
                    role=models.Role.manager, user_branch_id=b.id, is_active=True)
    db.add(u); db.commit()
    bid, uid = b.id, u.id
    login(client, seed["admin"])
    assert client.post(f"/admin/users/{uid}/delete",
                       follow_redirects=False).status_code == 302
    db.expire_all()
    assert db.get(models.User, uid) is None
    assert db.get(models.Branch, bid) is None


# ================= КПП — ko'rish va filial cheklovi =================
@pytest.fixture(scope="module")
def kpp_users(db, seed):
    """Bitta filialga biriktirilgan КПП va biriktirilmagan КПП."""
    from app import auth

    def mk(email, branches):
        u = models.User(full_name=email.split("@")[0], email=email,
                        hashed_password=auth.hash_password("test12345"),
                        role=models.Role.kpp, is_active=True)
        u.visible_branches = branches
        db.add(u)
        return u

    tied = mk("kpp1@t.uz", [seed["b1"]])
    free = mk("kpp_no_branch@t.uz", [])
    # shu testlar uchun toza taom — boshqa testlarda stopga tushmagan
    dish = models.MenuItem(name="Морс клюквенный", is_active=True)
    db.add(dish)
    db.commit()
    return {"tied": tied, "free": free, "dish": dish}


def test_kpp_sees_stoplist_page_and_filter(client, seed, kpp_users):
    """КПП stop-list sahifasiga kiradi va filtr paneli ko'rinadi."""
    login(client, kpp_users["tied"])
    r = client.get("/stoplist", follow_redirects=False)
    assert r.status_code == 200, "КПП stop-listni ko'rishi kerak"
    assert 'name="branch_id"' in r.text, "filtr КПП uchun ham bo'lishi kerak"
    assert 'name="reason"' in r.text


def test_kpp_filter_lists_only_assigned_branches(client, seed, kpp_users):
    """Filtrdagi filial ro'yxatida faqat biriktirilgan filial bo'ladi."""
    login(client, kpp_users["tied"])
    r = client.get("/stoplist")
    assert "Ресторан №12" in r.text
    assert "Ресторан №7" not in r.text


def test_kpp_does_not_see_other_branch_entries(client, seed, db, kpp_users):
    """Boshqa filial yozuvlari ro'yxatga ham, URL orqali ham tushmaydi."""
    login(client, seed["branch2"])                       # Ресторан №7 yozuv qo'shadi
    r = api_create(client, kpp_users["dish"].id)
    assert r.status_code == 201, r.text
    other_id = r.json()["created"][0]["id"]

    login(client, kpp_users["tied"])                     # КПП faqat Ресторан №12
    link = f"/stoplist/{other_id}"
    assert link not in client.get("/stoplist").text
    # begona filial ID si bilan filtr ham ma'lumot ochmaydi
    assert link not in client.get(f"/stoplist?branch_id={seed['b2'].id}").text
    # to'g'ridan-to'g'ri havola ham yopiq
    assert client.get(link, follow_redirects=False).status_code in (302, 403, 404)


def test_kpp_without_branches_sees_nothing(client, seed, kpp_users):
    """Filial biriktirilmagan КПП — sahifa ochiladi, lekin yozuv ko'rinmaydi."""
    login(client, kpp_users["free"])
    r = client.get("/stoplist", follow_redirects=False)
    assert r.status_code == 200
    assert "Ресторан №12" not in r.text and "Ресторан №7" not in r.text


def test_admin_can_grant_stoplist_to_any_role(client, seed, db):
    """Admin ruxsat bergan ijrochi ham stop-listni ko'radi (standartda ko'rmaydi)."""
    from app import auth
    u = models.User(full_name="exec", email="exec_stop@t.uz",
                    hashed_password=auth.hash_password("test12345"),
                    role=models.Role.executor, is_active=True)
    db.add(u); db.commit()

    login(client, u)
    assert client.get("/stoplist", follow_redirects=False).status_code == 302

    u.perms = '{"view_stop": true}'
    db.commit()
    login(client, u)
    assert client.get("/stoplist", follow_redirects=False).status_code == 200


def test_kpp_can_export_only_own_branches(client, seed, db, kpp_users):
    """КПП Excel yuklaydi, lekin faylda faqat o'z filiali yozuvlari bo'ladi."""
    import openpyxl

    # Ресторан №7 da yozuv bor (boshqa testdan) — КПП ga u ko'rinmasligi kerak
    login(client, kpp_users["tied"])
    r = client.get("/stoplist/export")
    assert r.status_code == 200, "КПП eksport qila olishi kerak"

    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    branches = {row[1] for row in wb.active.iter_rows(min_row=2, values_only=True)
                if row and row[1]}
    assert branches <= {"Ресторан №12"}, f"begona filial tushib qoldi: {branches}"


def test_kpp_without_branches_exports_empty(client, seed, kpp_users):
    """Filial biriktirilmagan КПП — fayl beriladi, lekin yozuvsiz."""
    import openpyxl

    login(client, kpp_users["free"])
    r = client.get("/stoplist/export")
    assert r.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    rows = [x for x in wb.active.iter_rows(min_row=2, values_only=True) if x and x[1]]
    assert rows == []
