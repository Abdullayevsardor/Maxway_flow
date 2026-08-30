"""MAXWAY — ishlarni saqlash va ijrochilarga yo'naltirish tizimi (FastAPI)."""
import os
import json
import hashlib
import urllib.request
import urllib.parse
import urllib.error
from datetime import datetime, timedelta
from typing import Optional, List

from fastapi import FastAPI, Depends, Request, Form, HTTPException, File, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from sqlalchemy import func, or_, text as sqltext
from sqlalchemy.orm import Session

from app.database import Base, engine, get_db, SessionLocal
from app import models, auth
from app.models import Role, Status, Priority

print(">>> [MAXWAY] create_all boshlandi", flush=True)
Base.metadata.create_all(bind=engine)
print(">>> [MAXWAY] create_all tugadi", flush=True)


def _ensure_columns():
    """Eski bazaga yetishmagan ustunlarни qo'shadi (SQLite + Postgres mos)."""
    from sqlalchemy import inspect as sa_inspect
    checks = [("attachments", "stage", "VARCHAR(10) DEFAULT 'request'"),
              ("requests", "subcategory_id", "INTEGER"),
              ("branches", "phone", "VARCHAR(40) DEFAULT ''"),
              ("branches", "director_name", "VARCHAR(120) DEFAULT ''"),
              ("branches", "tg_chat_ids", "TEXT DEFAULT ''"),
              ("notifications", "from_name", "VARCHAR(120)"),
              ("stop_entries", "resolved_at", "TIMESTAMP"),
              ("stop_entries", "supply_comment", "TEXT DEFAULT ''"),
              ("stop_entries", "supply_confirmed", "BOOLEAN DEFAULT FALSE"),
              ("stop_entries", "confirmed_by", "INTEGER"),
              ("stop_entries", "confirmed_at", "TIMESTAMP"),
              ("stop_entries", "updated_by", "INTEGER"),
              ("stop_entries", "updated_at", "TIMESTAMP"),
              ("users", "perms", "TEXT"),
              ("requests", "dep_number", "INTEGER")]
    try:
        insp = sa_inspect(engine)
        tables = insp.get_table_names()
        for table, col, ddl in checks:
            if table not in tables:
                continue
            cols = [c["name"] for c in insp.get_columns(table)]
            if col not in cols:
                with engine.begin() as conn:
                    conn.exec_driver_sql(
                        f"ALTER TABLE {table} ADD COLUMN {col} {ddl}")
                print(f">>> [MAXWAY] '{col}' ustuni '{table}' ga qo'shildi", flush=True)
    except Exception as e:
        print(">>> [MAXWAY] ensure_columns xato:", e, flush=True)


def _ensure_indexes():
    """Stop-list ro'yxati tez ochilishi uchun indekslar (SQLite + Postgres mos).
    create_all faqat yangi jadval yaratganda indeks qo'yadi — mavjud jadvalga qo'lda."""
    from sqlalchemy import inspect as sa_inspect
    idx = [("stop_entries", "ix_stop_entries_created_at", "created_at"),
           ("stop_entries", "ix_stop_entries_branch_id", "branch_id"),
           ("stop_entries", "ix_stop_entries_menu_item_id", "menu_item_id"),
           ("stop_entries", "ix_stop_entries_reason", "reason"),
           ("stop_entries", "ix_stop_entries_supply_confirmed", "supply_confirmed"),
           ("stop_entries", "ix_stop_entries_resolved", "resolved"),
           # ro'yxat doim (resolved + sana) bo'yicha o'qiladi — kompozit indeks
           ("stop_entries", "ix_stop_entries_resolved_created", "resolved, created_at")]
    try:
        insp = sa_inspect(engine)
        tables = insp.get_table_names()
        for table, name, cols in idx:
            if table not in tables:
                continue
            have = {i["name"] for i in insp.get_indexes(table)}
            if name in have:
                continue
            with engine.begin() as conn:
                conn.exec_driver_sql(
                    f"CREATE INDEX IF NOT EXISTS {name} ON {table} ({cols})")
            print(f">>> [MAXWAY] indeks '{name}' yaratildi", flush=True)
    except Exception as e:
        print(">>> [MAXWAY] ensure_indexes xato:", e, flush=True)


print(">>> [MAXWAY] ensure_columns boshlandi", flush=True)
_ensure_columns()
_ensure_indexes()
print(">>> [MAXWAY] ensure_columns tugadi", flush=True)


def _ensure_enum_values():
    """Postgres 'role' enum turiga yangi qiymatlarni (viewer) qo'shadi.
    Python enumga qiymat qo'shilsa, PG enum turi avtomatik yangilanmaydi."""
    try:
        if "postgres" not in str(engine.url):
            return  # SQLite'da enum matn sifatida — kerak emas
        from sqlalchemy import text
        conn = engine.connect().execution_options(isolation_level="AUTOCOMMIT")
        try:
            row = conn.execute(text(
                "SELECT t.typname FROM pg_type t JOIN pg_enum e ON e.enumtypid=t.oid "
                "WHERE e.enumlabel='executor' LIMIT 1")).first()
            if row:
                tname = row[0]
                for val in ("viewer", "kpp"):
                    conn.execute(text(f"ALTER TYPE {tname} ADD VALUE IF NOT EXISTS '{val}'"))
                print(f">>> [MAXWAY] enum '{tname}' ga viewer/kpp qo'shildi", flush=True)
        finally:
            conn.close()
    except Exception as e:
        print(">>> [MAXWAY] enum migratsiya xato:", e, flush=True)


_ensure_enum_values()


def _auto_seed():
    """Birinchi ishga tushishда (foydalanuvchи yo'q bo'lsa) demo ma'lumot yaratadi."""
    try:
        db = SessionLocal()
        has_user = db.query(models.User).first()
        db.close()
        if not has_user:
            import seed as _seed
            _seed.seed()
            print(">>> [MAXWAY] Demo ma'lumot yaratildi (auto-seed)", flush=True)
        else:
            print(">>> [MAXWAY] Foydalanuvchи bor — seed o'tkazib yuborildi", flush=True)
    except Exception as e:
        print(">>> [MAXWAY] auto-seed xato:", e, flush=True)


print(">>> [MAXWAY] auto_seed boshlandi", flush=True)
_auto_seed()
print(">>> [MAXWAY] auto_seed tugadi — ilova yaratilmoqda", flush=True)


def _admin_tools():
    """Startupда: barcha admin emaillarini ko'rsatadi va (ADMIN_RESET berilsa) parol tiklaydi.
    ADMIN_RESET formati: "email:yangiparol"  (masalan: a.ruzikulov@gmail.com:12345678)"""
    try:
        db = SessionLocal()
        # eski "На проверке" (on_check) zayavkalarni "В работе" (in_progress) ga o'tkazamiz
        try:
            n1 = db.query(models.Request).filter(models.Request.status == models.Status.on_check)\
                .update({models.Request.status: models.Status.in_progress}, synchronize_session=False)
            n2 = db.query(models.StatusHistory).filter(models.StatusHistory.status == models.Status.on_check)\
                .update({models.StatusHistory.status: models.Status.in_progress}, synchronize_session=False)
            if n1 or n2:
                db.commit()
                print(f">>> [MAXWAY] on_check ko'chirildi: zayavka={n1}, tarix={n2}", flush=True)
        except Exception as e:
            db.rollback()
            print(">>> [MAXWAY] on_check migratsiya xato:", e, flush=True)
        admins = db.query(models.User).filter(models.User.role == models.Role.admin).all()
        print(">>> [MAXWAY] Adminlar:", flush=True)
        for a in admins:
            print(f"      - {a.email}  (active={a.is_active})", flush=True)
        reset = os.environ.get("ADMIN_RESET", "").strip()
        if reset and ":" in reset:
            em, pw = reset.split(":", 1)
            em = em.lower().strip()
            u = db.query(models.User).filter(models.User.email == em).first()
            if u:
                u.hashed_password = auth.hash_password(pw)
                u.is_active = True
                db.commit()
                print(f">>> [MAXWAY] ✅ Parol tiklandi: {em}", flush=True)
            else:
                print(f">>> [MAXWAY] ⚠ ADMIN_RESET: '{em}' topilmadi", flush=True)
        db.close()
    except Exception as e:
        print(">>> [MAXWAY] admin_tools xato:", e, flush=True)


_admin_tools()

app = FastAPI(title="MAXWAY")
app.mount("/static", StaticFiles(directory="static"), name="static")


@app.middleware("http")
async def _no_cache(request: Request, call_next):
    """Sahifalar keshda saqlanmasin (eski qiymatlar ko'rinmasin). Statika keshlanaveradi."""
    response = await call_next(request)
    if not request.url.path.startswith("/static"):
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
    return response
templates = Jinja2Templates(directory="templates")
print(">>> [MAXWAY] ilova tayyor ✅", flush=True)

STATUS_LABELS = {
    "new": "Новая", "approved": "Одобрена", "in_progress": "В работе",
    "done": "Выполнена", "rejected": "Отклонена",
}
PRIORITY_LABELS = {"low": "Низкий", "medium": "Средний", "high": "Высокий"}
ROLE_LABELS = {"admin": "АДМИН", "manager": "МЕНЕДЖЕР", "executor": "ИСПОЛНИТЕЛЬ", "client": "ЗАКАЗЧИК", "viewer": "ПРОСМОТР", "kpp": "КПП"}
# Stop-list sabablari (spravochnik). Yangi yozuvlar shu ro'yxatdan tanlanadi.
REASON_LABELS = {
    "equipment_broken": "Сломалось оборудование",
    "supplier_no_product": "Нет продукта у поставщика",
    "wrong_order": "Неправильный заказ",
    "menu_removed": "Вывод из меню продукта",
    "sales_growth": "Рост продаж",
    "wrong_forecast": "Неправильный прогноз продаж",
    "supplier_late": "Поставщик опоздал",
    "supplier_stop": "На стопе у поставщика",
    "branch_no_order": "Не заказал филиал",
    "tech_problem": "Технический проблема",
}
CONFIRM_LABELS = {True: "ДА", False: "НЕТ"}
# Stop-list ro'yxati: sahifalash va saralash
STOP_PAGE_SIZES = (10, 25, 50, 100)
STOP_PAGE_SIZE_DEFAULT = 25
# tashqi nom -> saralanadigan ustun (SQL injection'ga yopiq: faqat shu kalitlar)
STOP_SORT_FIELDS = {
    "created_at": lambda: models.StopEntry.created_at,
    "branch": lambda: models.Branch.name,
    "dish": lambda: models.MenuItem.name,
    "reason": lambda: models.StopEntry.reason,
    "comment": lambda: models.StopEntry.comment,
    "supply_confirmed": lambda: models.StopEntry.supply_confirmed,
    "supply_comment": lambda: models.StopEntry.supply_comment,
    "resolved_at": lambda: models.StopEntry.resolved_at,
}
# matn maydonlari uchun chegaralar (backend validatsiyasi)
STOP_COMMENT_MAX = 1000

templates.env.globals.update(
    STATUS_LABELS=STATUS_LABELS, PRIORITY_LABELS=PRIORITY_LABELS,
    ROLE_LABELS=ROLE_LABELS, REASON_LABELS=REASON_LABELS,
    CONFIRM_LABELS=CONFIRM_LABELS, STOP_PAGE_SIZES=STOP_PAGE_SIZES,
    confirm_label=lambda v: CONFIRM_LABELS[bool(v)],
    APP_NAME="MAXWAY", now=datetime.utcnow,
)


def current_user(request: Request, db: Session):
    return auth.get_current_user_optional(request, db)


# ---------- Telegram bot ----------
def get_bot_token() -> str:
    """Token: MAXWAY_BOT_TOKEN env yoki telegram_token.txt fayldan."""
    tok = os.environ.get("MAXWAY_BOT_TOKEN", "").strip()
    if not tok:
        try:
            with open("telegram_token.txt", encoding="utf-8") as f:
                tok = f.read().strip()
        except FileNotFoundError:
            tok = ""
    # Agar "MAXWAY_BOT_TOKEN=xxx" ko'rinishida yozilgan bo'lsa, faqat qiymatini olamiz
    if "=" in tok:
        tok = tok.split("=", 1)[1].strip()
    return tok


def get_app_url() -> str:
    """Loyiha manzili: MAXWAY_URL env yoki maxway_url.txt, aks holда localhost."""
    u = os.environ.get("MAXWAY_URL", "").strip()
    if not u:
        try:
            with open("maxway_url.txt", encoding="utf-8") as f:
                u = f.read().strip()
        except FileNotFoundError:
            u = ""
    return u or "http://127.0.0.1:8000"


def get_stop_bot_token() -> str:
    """Stop-list xabarlari uchun ALOHIDA bot tokeni (ixtiyoriy).
    MAXWAY_STOP_BOT_TOKEN env yoki stop_bot_token.txt fayldan.
    Sozlanmagan bo'lsa — asosiy bot (MAXWAY_BOT_TOKEN) ishlatiladi."""
    tok = os.environ.get("MAXWAY_STOP_BOT_TOKEN", "").strip()
    if not tok:
        try:
            with open("stop_bot_token.txt", encoding="utf-8") as f:
                tok = f.read().strip()
        except FileNotFoundError:
            tok = ""
    if "=" in tok:
        tok = tok.split("=", 1)[1].strip()
    return tok or get_bot_token()


def send_telegram(chat_id: str, text: str, button_url: str = "",
                  button_text: str = "Открыть MAXWAY", token: str = ""):
    """Telegramга xabar yuboradi (ixtiyoriy tugma bilan). Xatolik tinch o'tadi.
    `token` berilmasa — asosiy bot tokeni ishlatiladi."""
    token = token or get_bot_token()
    if not token or not chat_id:
        return
    try:
        params = {"chat_id": chat_id, "text": text, "parse_mode": "HTML"}
        if button_url:
            params["reply_markup"] = json.dumps(
                {"inline_keyboard": [[{"text": button_text, "url": button_url}]]})
        url = f"https://api.telegram.org/bot{token}/sendMessage"
        data = urllib.parse.urlencode(params).encode()
        urllib.request.urlopen(urllib.request.Request(url, data=data), timeout=8)
    except Exception:
        pass


def branch_chat_ids(db: Session, r: models.Request, creator=None) -> List[str]:
    """Zayavka filialiga tegishli barcha telegram chat_id lar (takrorsiz).
    Manba: zayavka egasining chat_id si + filialga yozilgan qo'shimcha ID lar."""
    ids, seen = [], set()

    def _add(cid):
        cid = (cid or "").strip()
        if cid and cid not in seen:
            seen.add(cid)
            ids.append(cid)

    if creator is None and r.created_by:
        creator = db.get(models.User, r.created_by)
    if creator:
        _add(creator.telegram_chat_id)
    br_id = r.branch_id or (creator.user_branch_id if creator else None)
    if br_id:
        b = db.get(models.Branch, br_id)
        raw = (b.tg_chat_ids or "") if b else ""
        for part in raw.replace(";", ",").replace("\n", ",").replace(" ", ",").split(","):
            _add(part)
    return ids


def get_stop_channel() -> str:
    """Stop-list xabarlari uchun umumiy kanal/guruh ID si (ixtiyoriy).
    MAXWAY_STOP_CHANNEL env yoki stop_channel.txt fayldan. Bo'sh — kanal ishlatilmaydi."""
    ch = os.environ.get("MAXWAY_STOP_CHANNEL", "").strip()
    if not ch:
        try:
            with open("stop_channel.txt", encoding="utf-8") as f:
                ch = f.read().strip()
        except FileNotFoundError:
            ch = ""
    if "=" in ch:
        ch = ch.split("=", 1)[1].strip()
    return ch


def _send_async(chat_ids, text: str, button_url: str = "", token: str = ""):
    """Xabarlarni fon oqimida yuboradi — foydalanuvchi sahifasi kutib qolmasin."""
    ids = [c for c in chat_ids if c]
    if not ids:
        return
    import threading

    def _run():
        for cid in ids:
            send_telegram(cid, text, button_url=button_url, token=token)

    threading.Thread(target=_run, daemon=True).start()


def base_requests(db: Session, user):
    """Klient → o'z filiali; bo'limga biriktirilgan admin/menejer/ijrochi → faqat o'z bo'limi;
    bo'limsiz admin (категория = —) → barcha zayavkalar (bosh admin)."""
    q = db.query(models.Request)
    if user.role == Role.client:
        q = q.filter(models.Request.branch_id == user.user_branch_id)
    elif user.role == Role.kpp:
        ids = [b.id for b in user.visible_branches]
        q = q.filter(models.Request.branch_id.in_(ids if ids else [-1]))
    elif user.role in (Role.admin, Role.manager, Role.executor) and user.department_id:
        q = q.filter(models.Request.department_id == user.department_id)
    return q


def scoped_executors(db: Session, user):
    """Biriktirish uchun ijrochilar: bo'limga biriktirilgan admin/menejer faqat
    o'z bo'limidagi ijrochilarni ko'radi; bo'limsiz admin — hammasini."""
    q = db.query(models.User).filter(models.User.is_active == True)
    if user.role in (Role.admin, Role.manager, Role.executor) and user.department_id:
        q = q.filter(models.User.department_id == user.department_id)
    return q.all()


def scoped_departments(db: Session, user):
    """Bo'limga biriktirilgan admin/ijrochi faqat o'z bo'limini ko'radi."""
    q = db.query(models.Department)
    if user.role in (Role.admin, Role.manager, Role.executor) and user.department_id:
        q = q.filter(models.Department.id == user.department_id)
    return q.all()


def display_name(u):
    """Filial (client) bo'lsa — direktor ismi, aks holda full_name."""
    if u and u.role == Role.client and u.branch and u.branch.director_name:
        return u.branch.director_name
    return u.full_name if u else "—"


def add_history(db: Session, req: models.Request, status: models.Status, note=""):
    db.add(models.StatusHistory(request_id=req.id, status=status, note=note))


# ===================== AUTH =====================
@app.get("/", response_class=HTMLResponse)
def root(request: Request, db: Session = Depends(get_db)):
    return RedirectResponse("/dashboard" if current_user(request, db) else "/login", 302)


@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request, db: Session = Depends(get_db)):
    if current_user(request, db):
        return RedirectResponse("/dashboard", 302)
    return templates.TemplateResponse(request, "login.html", {"request": request})


@app.post("/login")
def login_submit(request: Request, email: str = Form(...), password: str = Form(...),
                 db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.email == email.lower()).first()
    if not user or not auth.verify_password(password, user.hashed_password):
        return templates.TemplateResponse(request, "login.html",
            {"request": request, "error": "Неверный email или пароль"}, status_code=401)
    token = auth.create_access_token(user.id)
    resp = RedirectResponse("/dashboard", 302)
    resp.set_cookie(auth.COOKIE_NAME, token, httponly=True, max_age=7200)
    return resp


@app.get("/register", response_class=HTMLResponse)
def register_page(request: Request, db: Session = Depends(get_db)):
    deps = db.query(models.Department).all()
    return templates.TemplateResponse(request, "register.html",
        {"request": request, "departments": deps})


@app.post("/register")
def register_submit(request: Request, full_name: str = Form(...), email: str = Form(...),
                    password: str = Form(...), department_id: Optional[int] = Form(None),
                    db: Session = Depends(get_db)):
    email = email.lower().strip()
    if db.query(models.User).filter(models.User.email == email).first():
        deps = db.query(models.Department).all()
        return templates.TemplateResponse(request, "register.html",
            {"request": request, "departments": deps,
             "error": "Bu email allaqachon ro'yxatdan o'tgan"}, status_code=400)
    user = models.User(full_name=full_name.strip(), email=email,
                       hashed_password=auth.hash_password(password),
                       role=Role.executor, department_id=department_id)
    db.add(user); db.commit(); db.refresh(user)
    token = auth.create_access_token(user.id)
    resp = RedirectResponse("/dashboard", 302)
    resp.set_cookie(auth.COOKIE_NAME, token, httponly=True, max_age=7200)
    return resp


@app.get("/logout")
def logout():
    resp = RedirectResponse("/login", 302)
    resp.delete_cookie(auth.COOKIE_NAME)
    return resp


# ===================== DASHBOARD =====================
@app.get("/dashboard", response_class=HTMLResponse)
def dashboard(request: Request, db: Session = Depends(get_db),
              department_id: str = "", subcategory_id: str = "", assignee: str = "",
              customer: str = "", date_from: str = "", date_to: str = "", unassigned: str = "",
              month: str = "", branch_id: str = ""):
    user = current_user(request, db)
    if not user:
        return RedirectResponse("/login", 302)
    q = base_requests(db, user)
    from sqlalchemy import func as _func
    dep_counts = dict(q.with_entities(models.Request.department_id, _func.count(models.Request.id))
                      .group_by(models.Request.department_id).all())
    # КПП uchun: zayavkalar filial va kategoriya bo'yicha guruhlangan
    kpp_groups = None
    if user.role == Role.kpp:
        tmp = {}
        for r in q.order_by(models.Request.created_at.desc()).all():
            bn = r.branch_obj.name if r.branch_obj else (r.branch or "—")
            cn = r.department.name if r.department else "—"
            tmp.setdefault(bn, {}).setdefault(cn, []).append(r)
        kpp_groups = [{"branch": bn,
                       "cats": [{"cat": cn, "reqs": tmp[bn][cn]} for cn in sorted(tmp[bn])]}
                      for bn in sorted(tmp)]

    # ---- Filtrlar (dashboardda) ----
    f_dep = int(department_id) if str(department_id).isdigit() else None
    f_sub = int(subcategory_id) if str(subcategory_id).isdigit() else None
    f_asg = int(assignee) if str(assignee).isdigit() else None
    f_branch = int(branch_id) if str(branch_id).isdigit() else None
    fq = base_requests(db, user)
    if f_dep:
        fq = fq.filter(models.Request.department_id == f_dep)
    if f_sub:
        fq = fq.filter(models.Request.subcategory_id == f_sub)
    if f_asg:
        fq = fq.filter(models.Request.assignees.any(models.User.id == f_asg))
    if f_branch:
        fq = fq.filter(models.Request.branch_id == f_branch)
    if customer.strip():
        fq = fq.filter(models.Request.customer_name.ilike(f"%{customer.strip()}%"))
    if unassigned:
        fq = fq.filter(~models.Request.assignees.any(),
                       models.Request.status.notin_([Status.done, Status.rejected]))
    if month.strip():
        try:
            m0 = datetime.strptime(month.strip(), "%Y-%m")
            m1 = m0.replace(year=m0.year + 1, month=1) if m0.month == 12 else m0.replace(month=m0.month + 1)
            fq = fq.filter(models.Request.created_at >= m0, models.Request.created_at < m1)
        except ValueError:
            pass
    if date_from.strip():
        try:
            fq = fq.filter(models.Request.created_at >= datetime.strptime(date_from.strip(), "%Y-%m-%d"))
        except ValueError:
            pass
    if date_to.strip():
        try:
            fq = fq.filter(models.Request.created_at < datetime.strptime(date_to.strip(), "%Y-%m-%d") + timedelta(days=1))
        except ValueError:
            pass
    filtered = fq.order_by(models.Request.created_at.desc()).all()
    any_filter = bool(f_dep or f_sub or f_asg or f_branch or customer.strip() or unassigned or month.strip() or date_from.strip() or date_to.strip())

    # bo'lim -> podkategoriyalar (dinamik filtr uchun)
    subcats_map = {}
    for d in scoped_departments(db, user):
        subcats_map[d.id] = [{"id": s.id, "name": s.name} for s in d.subcategories]
    # zakazchiklar ro'yxati (datalist uchun)
    customers = [c[0] for c in base_requests(db, user).with_entities(models.Request.customer_name)
                 .filter(models.Request.customer_name != "").distinct().all() if c[0]]
    _today = datetime.utcnow() + timedelta(hours=5)
    month_start = _today.replace(day=1).strftime("%Y-%m-%d")
    today_str = _today.strftime("%Y-%m-%d")
    cur_month = _today.strftime("%Y-%m")

    ctx = {
        "request": request, "user": user, "active": "dashboard",
        "total": q.count(),
        "new": q.filter(models.Request.status == Status.new).count(),
        "in_progress": q.filter(models.Request.status == Status.in_progress).count(),
        "done": q.filter(models.Request.status == Status.done).count(),
        "unassigned": q.filter(~models.Request.assignees.any(),
                               ~models.Request.status.in_([Status.done, Status.rejected])).count(),
        "departments": scoped_departments(db, user),
        "dep_counts": dep_counts,
        "kpp_groups": kpp_groups,
        "recent": q.order_by(models.Request.created_at.desc()).limit(10).all(),
        # filtr
        "filtered": filtered, "any_filter": any_filter,
        "executors": scoped_executors(db, user),
        "subcats_map": subcats_map,
        "f_dep": f_dep, "f_sub": f_sub, "f_asg": f_asg,
        "f_customer": customer, "f_date_from": date_from, "f_date_to": date_to,
        "f_unassigned": unassigned, "f_month": month,
        "customers": sorted(customers), "month_start": month_start, "today_str": today_str,
        "cur_month": cur_month,
        "branches": db.query(models.Branch).order_by(models.Branch.name).all(),
        "f_branch": f_branch,
    }
    return templates.TemplateResponse(request, "dashboard.html", ctx)


# ===================== ZAYAVKALAR =====================
@app.get("/requests", response_class=HTMLResponse)
def requests_page(request: Request, department_id: str = "",
                  status: Optional[str] = None, q: Optional[str] = None,
                  subcategory_id: str = "", date_from: str = "", date_to: str = "",
                  customer: str = "", assignee: str = "",
                  unassigned: str = "",
                  db: Session = Depends(get_db)):
    user = current_user(request, db)
    if not user:
        return RedirectResponse("/login", 302)
    # bo'sh qatorlarni xavfsiz int/None ga aylantiramiz
    department_id = int(department_id) if str(department_id).isdigit() else None
    subcategory_id = int(subcategory_id) if str(subcategory_id).isdigit() else None
    assignee = int(assignee) if str(assignee).isdigit() else None
    unassigned = 1 if unassigned else None
    query = base_requests(db, user)
    if department_id:
        query = query.filter(models.Request.department_id == department_id)
    if subcategory_id:
        query = query.filter(models.Request.subcategory_id == subcategory_id)
    if unassigned:
        query = query.filter(~models.Request.assignees.any(),
                             models.Request.status.notin_([Status.done, Status.rejected]))
    if status == "overdue":
        query = query.filter(models.Request.deadline.isnot(None),
                             models.Request.deadline < datetime.utcnow(),
                             models.Request.status.notin_([Status.done, Status.rejected]))
    elif status in STATUS_LABELS:
        query = query.filter(models.Request.status == status)
    if q:
        query = query.filter(models.Request.title.ilike(f"%{q}%"))
    if customer.strip():
        query = query.filter(models.Request.customer_name.ilike(f"%{customer.strip()}%"))
    if assignee:
        query = query.filter(models.Request.assignees.any(models.User.id == assignee))
    for fmt in ("%Y-%m-%d",):
        if date_from.strip():
            try:
                query = query.filter(models.Request.created_at >= datetime.strptime(date_from.strip(), fmt))
            except ValueError:
                pass
        if date_to.strip():
            try:
                query = query.filter(models.Request.created_at < datetime.strptime(date_to.strip(), fmt) + timedelta(days=1))
            except ValueError:
                pass
    items = query.order_by(models.Request.created_at.desc()).all()

    base = base_requests(db, user)
    counts = {
        "all": base.count(),
        "new": base.filter(models.Request.status == Status.new).count(),
        "approved": base.filter(models.Request.status == Status.approved).count(),
        "in_progress": base.filter(models.Request.status == Status.in_progress).count(),
        "on_check": base.filter(models.Request.status == Status.on_check).count(),
        "done": base.filter(models.Request.status == Status.done).count(),
        "rejected": base.filter(models.Request.status == Status.rejected).count(),
        "overdue": base.filter(models.Request.deadline.isnot(None),
                              models.Request.deadline < datetime.utcnow(),
                              models.Request.status.notin_([Status.done, Status.rejected])).count(),
    }
    selected_dep = db.get(models.Department, department_id) if department_id else None
    subcats = db.query(models.Subcategory).order_by(models.Subcategory.name).all()
    subcats_map = {}
    for s in subcats:
        subcats_map.setdefault(s.department_id, []).append({"id": s.id, "name": s.name})
    return templates.TemplateResponse(request, "requests.html", {
        "request": request, "user": user, "active": "requests",
        "items": items, "departments": scoped_departments(db, user),
        "executors": scoped_executors(db, user), "selected_dep": selected_dep,
        "selected_status": status, "counts": counts, "search": q or "",
        "branches": db.query(models.Branch).order_by(models.Branch.name).all(),
        "subcats_map": subcats_map, "all_subcats": subcats,
        "f_subcategory": subcategory_id, "f_date_from": date_from, "f_date_to": date_to,
        "f_customer": customer, "f_assignee": assignee, "f_unassigned": unassigned,
    })


@app.get("/requests/{req_id}", response_class=HTMLResponse)
def request_detail(req_id: int, request: Request, db: Session = Depends(get_db)):
    user = current_user(request, db)
    if not user:
        return RedirectResponse("/login", 302)
    r = db.get(models.Request, req_id)
    if not r:
        raise HTTPException(404, "Заявка не найдена")
    if user.role == Role.client and r.branch_id != user.user_branch_id:
        return RedirectResponse("/requests", 302)
    # bo'limga biriktirilgan admin/menejer/ijrochi boshqa bo'lim zayavkasini ko'rolmaydi
    if user.role in (Role.admin, Role.manager, Role.executor) and user.department_id \
            and r.department_id != user.department_id:
        return RedirectResponse("/requests", 302)
    return templates.TemplateResponse(request, "request_detail.html", {
        "request": request, "user": user, "active": "requests", "r": r,
        "executors": scoped_executors(db, user),
    })


@app.post("/requests/create")
def create_request(request: Request, title: str = Form(...), description: str = Form(""),
                   department_id: int = Form(...), subcategory_id: Optional[int] = Form(None),
                   priority: str = Form("medium"),
                   customer_name: str = Form(""), customer_email: str = Form(""),
                   customer_phone: str = Form(""), branch_id: Optional[int] = Form(None),
                   deadline: str = Form(""), photos: List[UploadFile] = File([]),
                   videos: List[UploadFile] = File([]), db: Session = Depends(get_db)):
    user = current_user(request, db)
    if not user:
        return RedirectResponse("/login", 302)
    if not has_perm(user, "create_request"):
        return RedirectResponse("/requests", 302)
    if not title.strip():
        return RedirectResponse("/requests?err=title", 302)
    dl = None
    if deadline:
        for fmt in ("%Y-%m-%dT%H:%M", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"):
            try:
                dl = datetime.strptime(deadline, fmt)
                break
            except ValueError:
                dl = None
    branch_name = ""
    branch_phone = ""
    branch_director = ""
    if user.role == Role.client and user.user_branch_id:
        branch_id = user.user_branch_id
    if branch_id:
        b = db.get(models.Branch, branch_id)
        if b:
            branch_name = b.name
            branch_phone = b.phone or ""
            branch_director = b.director_name or ""
    # Заказчик: filial bo'lsa — direktor ismi; boshqalar — profil ismi
    if user.role == Role.client:
        cust_name = customer_name.strip() or branch_director or user.full_name or branch_name
        cust_phone = customer_phone.strip() or branch_phone or (user.phone or "")
    else:
        cust_name = customer_name.strip() or user.full_name
        cust_phone = customer_phone.strip() or (user.phone or "")
    cust_email = customer_email.strip() or user.email
    # bo'lim ichidagi tartib raqami (1 dan boshlanadi, har bo'lim alohida)
    from sqlalchemy import func as _f
    last_no = db.query(_f.max(models.Request.dep_number)).filter(
        models.Request.department_id == department_id).scalar() or 0
    r = models.Request(title=title.strip(), description=description.strip(),
                       department_id=department_id,
                       subcategory_id=subcategory_id or None,
                       priority=Priority(priority),
                       status=Status.new, created_by=user.id,
                       dep_number=last_no + 1,
                       customer_name=cust_name, customer_email=cust_email,
                       customer_phone=cust_phone, branch=branch_name,
                       branch_id=branch_id, deadline=dl)
    db.add(r); db.flush()
    add_history(db, r, Status.new, "Заявка создана")

    img_ext = (".jpg", ".jpeg", ".png", ".webp", ".gif")
    vid_ext = (".mp4", ".mov", ".webm", ".avi", ".mkv", ".m4v", ".3gp")
    os.makedirs("static/uploads/requests", exist_ok=True)
    idx = 0
    for up in list(photos) + list(videos):
        if not up or not up.filename:
            continue
        ext = os.path.splitext(up.filename)[1].lower()
        if ext in img_ext:
            kind = "image"
        elif ext in vid_ext:
            kind = "video"
        else:
            continue
        fname = f"req_{r.id}_{idx}{ext}"
        idx += 1
        with open(os.path.join("static", "uploads", "requests", fname), "wb") as fh:
            fh.write(up.file.read())
        db.add(models.Attachment(request_id=r.id,
                                 file_path=f"/static/uploads/requests/{fname}", kind=kind))
    db.commit()
    notify_category(db, r)
    return RedirectResponse(f"/requests/{r.id}", 302)


def notify_category(db: Session, r: models.Request):
    """Zayavка kategoriyasidagi xodimlarга va adminларга bildirishnoma (sayt + Telegram)."""
    dep = r.department
    recipients = db.query(models.User).filter(
        models.User.is_active == True,
        models.User.id != r.created_by,
        models.User.role.in_([Role.executor, Role.manager, Role.admin, Role.viewer]),
        ((models.User.department_id == r.department_id)
         | ((models.User.role == Role.admin) & (models.User.department_id.is_(None)))
         | (models.User.role == Role.viewer))
    ).all()
    seen = set()
    link = f"/requests/{r.id}"
    tg_text = (f"🔔 <b>Новая заявка — MAXWAY</b>\n\n"
               f"📌 <b>{r.title}</b>\n"
               f"🏷 Категория: {dep.name if dep else '—'}\n"
               f"⚡️ Приоритет: {PRIORITY_LABELS.get(r.priority.value)}\n"
               f"🏢 Филиал: {r.branch_obj.name if r.branch_obj else (r.branch or '—')}\n"
               f"👤 Заказчик: {r.customer_name or '—'}\n"
               f"📞 Телефон: {r.customer_phone or '—'}\n"
               f"⏰ Дедлайн: {r.deadline.strftime('%d.%m.%Y') if r.deadline else '—'}")
    site_text = f"Новая заявка «{r.title}» — {dep.name if dep else 'без категории'}"
    for u in recipients:
        if u.id in seen:
            continue
        seen.add(u.id)
        db.add(models.Notification(user_id=u.id, text=site_text, link=link,
                                   from_name=r.customer_name or "—"))
        if u.telegram_chat_id:
            send_telegram(u.telegram_chat_id, tg_text,
                          button_url=f"{get_app_url()}{link}")
    db.commit()


@app.get("/api/notifications")
def api_notifications(request: Request, db: Session = Depends(get_db)):
    user = current_user(request, db)
    if not user:
        return JSONResponse({"unread": 0, "items": []})
    items = db.query(models.Notification).filter(models.Notification.user_id == user.id)\
        .order_by(models.Notification.created_at.desc()).limit(50).all()
    # bo'limga biriktirilgan foydalanuvchi faqat o'z bo'limi zayavkalariga oid xabarlarni ko'radi
    if user.role in (Role.admin, Role.manager, Role.executor) and user.department_id:
        kept = []
        for n in items:
            ok = True
            if n.link and n.link.startswith("/requests/"):
                try:
                    rid = int(n.link.rsplit("/", 1)[1])
                    req = db.get(models.Request, rid)
                    if req and req.department_id != user.department_id:
                        ok = False
                except (ValueError, IndexError):
                    pass
            if ok:
                kept.append(n)
        items = kept[:20]
    else:
        items = items[:20]
    unread = sum(1 for n in items if not n.is_read)
    return JSONResponse({
        "unread": unread,
        "items": [{"text": n.text, "link": n.link, "is_read": n.is_read,
                   "from_name": n.from_name or "",
                   "time": n.created_at.strftime("%d.%m %H:%M")} for n in items]
    })


@app.post("/api/notifications/read")
def api_notifications_read(request: Request, db: Session = Depends(get_db)):
    user = current_user(request, db)
    if user:
        db.query(models.Notification).filter(
            models.Notification.user_id == user.id,
            models.Notification.is_read == False).update({models.Notification.is_read: True})
        db.commit()
    return JSONResponse({"ok": True})


@app.post("/requests/{req_id}/assign")
def assign_request(req_id: int, request: Request, assigned_to: int = Form(...),
                   db: Session = Depends(get_db)):
    user = current_user(request, db)
    if not user:
        return RedirectResponse("/login", 302)
    if not has_perm(user, "assign"):
        return RedirectResponse(f"/requests/{req_id}", 302)
    r = db.get(models.Request, req_id)
    if r:
        # bo'limga biriktirilgan admin/menejer faqat o'z bo'limida ish qiladi
        if user.role in (Role.admin, Role.manager, Role.executor) and user.department_id:
            if r.department_id != user.department_id:
                return RedirectResponse("/requests", 302)
            assignee_chk = db.get(models.User, assigned_to)
            if not assignee_chk or assignee_chk.department_id != user.department_id:
                return RedirectResponse(f"/requests/{req_id}", 302)
        assignee = db.get(models.User, assigned_to)
        if not assignee:
            return RedirectResponse(f"/requests/{req_id}", 302)
        # ro'yxatga qo'shamiz (avvalgilarni almashtirmaymiz)
        if assignee not in r.assignees:
            r.assignees.append(assignee)
        r.assigned_to = assigned_to   # legacy: oxirgi biriktirilgan
        # statusni majburan o'zgartirmaymiz — ijrochi o'zi: Одобрить → Начать работу
        if r.status == Status.rejected:
            r.status = Status.new
        add_history(db, r, r.status, f"Исполнитель назначен: {assignee.full_name}")
        db.commit()
        # Telegram xabari (ijrochining chat_id si bo'lsa)
        if assignee.telegram_chat_id:
            br = r.branch_obj.name if r.branch_obj else (r.branch or "—")
            text = (f"🔔 <b>Новое назначение — MAXWAY</b>\n\n"
                    f"📌 <b>{r.title}</b>\n"
                    f"🏷 Исполнитель: {assignee.full_name}\n"
                    f"⚡️ Приоритет: {PRIORITY_LABELS.get(r.priority.value)}\n"
                    f"🏢 Филиал: {br}\n"
                    f"👤 Заказчик: {r.customer_name or '—'}\n"
                    f"📞 Телефон: {r.customer_phone or '—'}\n"
                    f"⏰ Дедлайн: {r.deadline.strftime('%d.%m.%Y') if r.deadline else '—'}")
            send_telegram(assignee.telegram_chat_id, text,
                          button_url=f"{get_app_url()}/requests/{r.id}")
    return RedirectResponse(f"/requests/{req_id}", 302)


@app.post("/requests/{req_id}/unassign/{uid}")
def unassign_request(req_id: int, uid: int, request: Request, db: Session = Depends(get_db)):
    user = current_user(request, db)
    if not user or user.role not in (Role.admin, Role.manager):
        return RedirectResponse("/login", 302)
    r = db.get(models.Request, req_id)
    p = db.get(models.User, uid)
    if r and p and p in r.assignees:
        r.assignees.remove(p)
        # legacy assigned_to ni yangilaymiz
        r.assigned_to = r.assignees[-1].id if r.assignees else None
        add_history(db, r, r.status, f"Исполнитель снят: {p.full_name}")
        db.commit()
    return RedirectResponse(f"/requests/{req_id}", 302)


@app.post("/requests/{req_id}/reject")
def reject_request(req_id: int, request: Request, reason: str = Form(""),
                   db: Session = Depends(get_db)):
    user = current_user(request, db)
    if not user:
        return RedirectResponse("/login", 302)
    r = db.get(models.Request, req_id)
    if r:
        r.status = Status.rejected
        r.reject_reason = reason.strip()
        add_history(db, r, Status.rejected, reason.strip()[:200])
        db.commit()
    return RedirectResponse(f"/requests/{req_id}", 302)


@app.post("/requests/{req_id}/delete")
def delete_request(req_id: int, request: Request, db: Session = Depends(get_db)):
    """Заявкани butunlay o'chiradi — admin yoki menejer."""
    user = current_user(request, db)
    if not user or user.role not in (Role.admin, Role.manager):
        return RedirectResponse("/login", 302)
    r = db.get(models.Request, req_id)
    if r:
        # diskдаги fayllarни ham o'chiramiz
        for att in r.attachments:
            try:
                fp = att.file_path.lstrip("/")
                if os.path.exists(fp):
                    os.remove(fp)
            except Exception:
                pass
        # shu zayavkaga tegishli bildirishnomalarni o'chiramiz (link orqali bog'langan)
        db.query(models.Notification).filter(
            models.Notification.link == f"/requests/{req_id}").delete(synchronize_session=False)
        db.delete(r)   # comments/history/attachments cascade bilan o'chadi
        db.commit()
    return RedirectResponse("/requests", 302)


@app.post("/requests/{req_id}/status")
def change_status(req_id: int, request: Request, status: str = Form(...),
                  db: Session = Depends(get_db)):
    user = current_user(request, db)
    if not user:
        return RedirectResponse("/login", 302)
    r = db.get(models.Request, req_id)
    if r and status in STATUS_LABELS:
        notes = {
            "new": "Статус изменён",
            "approved": "Одобрено исполнителем",
            "in_progress": "Работа начата",
            "on_check": "Отправлено на проверку",
            "done": "Работа завершена",
            "rejected": "Отклонено",
        }
        r.status = Status(status)
        # ijrochi ishni boshlasa va hali biriktirilmagan bo'lsa — o'zi ro'yxatga qo'shiladi
        if user.role == Role.executor and not r.assignees \
                and status in ("approved", "in_progress", "on_check", "done"):
            r.assignees.append(user)
            r.assigned_to = user.id
        add_history(db, r, Status(status), notes.get(status, "Статус изменён"))
        db.commit()
    ref = request.headers.get("referer", f"/requests/{req_id}")
    return RedirectResponse(ref, 302)


@app.post("/requests/{req_id}/comment")
def add_comment(req_id: int, request: Request, text: str = Form(...),
                db: Session = Depends(get_db)):
    user = current_user(request, db)
    if not user or user.role in (Role.viewer, Role.kpp):
        return RedirectResponse("/login", 302)
    r = db.get(models.Request, req_id)
    if r and text.strip():
        db.add(models.Comment(request_id=req_id, user_id=user.id, text=text.strip()))
        db.commit()
        # zayavка egasiga bildirishnoma (boshqa odam izoh yozsa)
        if r.created_by and r.created_by != user.id:
            link = f"/requests/{r.id}"
            db.add(models.Notification(user_id=r.created_by,
                   text=f"Новый комментарий к «{r.title}»: {text.strip()[:60]}", link=link,
                   from_name=display_name(user)))
            db.commit()
            creator = db.get(models.User, r.created_by)
            tg_text = (f"💬 <b>Новый комментарий — MAXWAY</b>\n\n"
                       f"📌 <b>{r.title}</b>\n"
                       f"👤 {user.full_name}:\n{text.strip()}")
            for cid in branch_chat_ids(db, r, creator):
                send_telegram(cid, tg_text, button_url=f"{get_app_url()}{link}")
    return RedirectResponse(f"/requests/{req_id}", 302)


@app.post("/requests/{req_id}/solution")
def add_solution(req_id: int, request: Request, comment: str = Form(""),
                 photos: List[UploadFile] = File([]), videos: List[UploadFile] = File([]),
                 db: Session = Depends(get_db)):
    """Ijrochи yechimни (rasm/video/izoh) biriktiradi."""
    user = current_user(request, db)
    if not user or user.role == Role.client:
        return RedirectResponse("/login", 302)
    r = db.get(models.Request, req_id)
    if not r:
        return RedirectResponse("/requests", 302)
    img_ext = (".jpg", ".jpeg", ".png", ".webp", ".gif")
    vid_ext = (".mp4", ".mov", ".webm", ".avi", ".mkv", ".m4v", ".3gp")
    # kamida bittasi bo'lishi shart: rasm YOKI video YOKI izoh
    has_file = any(up and up.filename for up in (list(photos) + list(videos)))
    if not has_file and not comment.strip():
        return RedirectResponse(f"/requests/{req_id}?err=solution", 302)
    os.makedirs("static/uploads/requests", exist_ok=True)
    idx = db.query(models.Attachment).filter(models.Attachment.request_id == r.id).count()
    for up in list(photos) + list(videos):
        if not up or not up.filename:
            continue
        ext = os.path.splitext(up.filename)[1].lower()
        kind = "image" if ext in img_ext else ("video" if ext in vid_ext else None)
        if not kind:
            continue
        fname = f"req_{r.id}_sol_{idx}{ext}"
        idx += 1
        with open(os.path.join("static", "uploads", "requests", fname), "wb") as fh:
            fh.write(up.file.read())
        db.add(models.Attachment(request_id=r.id,
               file_path=f"/static/uploads/requests/{fname}", kind=kind, stage="solution"))
    if comment.strip():
        db.add(models.Comment(request_id=r.id, user_id=user.id,
               text="✅ Решение: " + comment.strip()))
    # yechim yuborilgach — ish yakunlandi
    if r.status not in (Status.done, Status.rejected):
        r.status = Status.done
        add_history(db, r, Status.done, "Работа завершена")
    db.commit()
    # zayavка egasiga (klиентга) bildirishnoma
    if r.created_by and r.created_by != user.id:
        link = f"/requests/{r.id}"
        db.add(models.Notification(user_id=r.created_by,
               text=f"Добавлено решение по заявке «{r.title}»", link=link,
               from_name=display_name(user)))
        db.commit()
        creator = db.get(models.User, r.created_by)
        tg_text = (f"✅ <b>Решение добавлено — MAXWAY</b>\n\n📌 <b>{r.title}</b>\n"
                   f"👤 {user.full_name}"
                   + (f"\n💬 {comment.strip()}" if comment.strip() else ""))
        for cid in branch_chat_ids(db, r, creator):
            send_telegram(cid, tg_text, button_url=f"{get_app_url()}{link}")
    return RedirectResponse(f"/requests/{req_id}", 302)


@app.post("/comments/{cid}/delete")
def delete_comment(cid: int, request: Request, db: Session = Depends(get_db)):
    user = current_user(request, db)
    if not user:
        return RedirectResponse("/login", 302)
    c = db.get(models.Comment, cid)
    if c and (user.role in (Role.admin, Role.manager) or c.user_id == user.id):
        rid = c.request_id
        db.delete(c); db.commit()
        return RedirectResponse(f"/requests/{rid}", 302)
    return RedirectResponse("/requests", 302)


# ===================== PROFIL =====================
@app.get("/profile", response_class=HTMLResponse)
def profile_page(request: Request, db: Session = Depends(get_db)):
    user = current_user(request, db)
    if not user:
        return RedirectResponse("/login", 302)
    return templates.TemplateResponse(request, "profile.html",
        {"request": request, "user": user, "active": "profile", "saved": False})


@app.post("/profile/update")
def profile_update(request: Request, full_name: str = Form(...), phone: str = Form(""),
                   bio: str = Form(""), telegram_chat_id: str = Form(""),
                   email: str = Form(""),
                   current_password: str = Form(""),
                   new_password: str = Form(""), confirm_password: str = Form(""),
                   photo: UploadFile = File(None), db: Session = Depends(get_db)):
    user = current_user(request, db)
    if not user:
        return RedirectResponse("/login", 302)
    # filial (client) bo'lsa — ism/telefon filial direktoriga yoziladi (filial nomi o'zgarmaydi)
    if user.role == Role.client and user.user_branch_id:
        b = db.get(models.Branch, user.user_branch_id)
        if b:
            b.director_name = full_name.strip()
            b.phone = phone.strip()
    else:
        user.full_name = full_name.strip()
        user.phone = phone.strip()
        # email (login) — faqat ADMIN o'zgartira oladi
        new_email = email.lower().strip()
        if user.role == Role.admin and new_email and new_email != user.email:
            taken = db.query(models.User).filter(
                models.User.email == new_email, models.User.id != user.id).first()
            if not taken:
                user.email = new_email
    user.bio = bio.strip()
    user.telegram_chat_id = telegram_chat_id.strip() or None
    # rasm yuklash
    if photo is not None and photo.filename:
        ext = os.path.splitext(photo.filename)[1].lower()
        if ext in (".jpg", ".jpeg", ".png", ".webp", ".gif"):
            os.makedirs("static/uploads", exist_ok=True)
            fname = f"avatar_{user.id}{ext}"
            with open(os.path.join("static", "uploads", fname), "wb") as f:
                f.write(photo.file.read())
            user.avatar = f"/static/uploads/{fname}"
    # parol o'zgartirish (yangi == tasdiq va joriy to'g'ri bo'lsa)
    msg = None
    if new_password:
        if new_password != confirm_password:
            msg = "Новые пароли не совпадают"
        elif not auth.verify_password(current_password, user.hashed_password):
            msg = "Текущий пароль неверный"
        else:
            user.hashed_password = auth.hash_password(new_password)
    db.commit()
    return templates.TemplateResponse(request, "profile.html",
        {"request": request, "user": user, "active": "profile",
         "saved": True, "msg": msg})


# ===================== IJROCHILAR =====================
@app.get("/executors", response_class=HTMLResponse)
def executors_page(request: Request, db: Session = Depends(get_db)):
    user = current_user(request, db)
    if not user:
        return RedirectResponse("/login", 302)
    people_q = db.query(models.User)
    # bo'limga biriktirilgan admin/ijrochi faqat o'z bo'limidagilarni ko'radi
    if user.role in (Role.admin, Role.manager, Role.executor) and user.department_id:
        people_q = people_q.filter(models.User.department_id == user.department_id)
    people = people_q.order_by(models.User.full_name).all()
    stats = {}
    for p in people:
        rq = db.query(models.Request).filter(models.Request.assignees.any(models.User.id == p.id))
        stats[p.id] = {
            "total": rq.count(),
            "in_progress": rq.filter(models.Request.status == Status.in_progress).count(),
            "done": rq.filter(models.Request.status == Status.done).count(),
        }
    return templates.TemplateResponse(request, "executors.html",
        {"request": request, "user": user, "active": "executors",
         "people": people, "stats": stats})


@app.post("/executors/{uid}/toggle")
def toggle_executor(uid: int, request: Request, db: Session = Depends(get_db)):
    user = current_user(request, db)
    if not user:
        return RedirectResponse("/login", 302)
    if user.role not in (Role.admin, Role.manager):
        return RedirectResponse("/executors", 302)
    p = db.get(models.User, uid)
    if p:
        p.is_active = not p.is_active
        db.commit()
    return RedirectResponse("/executors", 302)


@app.post("/executors/create")
def create_executor(request: Request, full_name: str = Form(...), email: str = Form(...),
                    password: str = Form("12345678"), position: str = Form(""),
                    specialization: str = Form(""), phone: str = Form(""),
                    db: Session = Depends(get_db)):
    user = current_user(request, db)
    if not user:
        return RedirectResponse("/login", 302)
    if user.role not in (Role.admin, Role.manager):
        return RedirectResponse("/executors", 302)
    email = email.lower().strip()
    if not db.query(models.User).filter(models.User.email == email).first():
        db.add(models.User(full_name=full_name.strip(), email=email,
                           hashed_password=auth.hash_password(password or "12345678"),
                           role=Role.executor, position=position.strip(),
                           specialization=specialization.strip(), phone=phone.strip(),
                           department_id=user.department_id, is_active=True))
        db.commit()
    return RedirectResponse("/executors", 302)


@app.post("/executors/{uid}/delete")
def delete_executor(uid: int, request: Request, db: Session = Depends(get_db)):
    user = current_user(request, db)
    if not user:
        return RedirectResponse("/login", 302)
    if user.role not in (Role.admin, Role.manager):
        return RedirectResponse("/executors", 302)
    p = db.get(models.User, uid)
    if p and p.id != user.id:
        db.query(models.Request).filter(models.Request.assigned_to == uid)\
            .update({models.Request.assigned_to: None})
        db.delete(p); db.commit()
    return RedirectResponse("/executors", 302)


@app.post("/executors/{uid}/edit")
def edit_executor(uid: int, request: Request, full_name: str = Form(...),
                  position: str = Form(""), specialization: str = Form(""),
                  phone: str = Form(""), schedule: str = Form(""),
                  experience: str = Form(""), telegram: str = Form(""),
                  db: Session = Depends(get_db)):
    user = current_user(request, db)
    if not user:
        return RedirectResponse("/login", 302)
    if user.role not in (Role.admin, Role.manager):
        return RedirectResponse("/executors", 302)
    p = db.get(models.User, uid)
    if p:
        p.full_name = full_name.strip()
        p.position = position.strip()
        p.specialization = specialization.strip()
        p.phone = phone.strip()
        p.schedule = schedule.strip()
        p.experience = experience.strip()
        p.telegram = telegram.strip()
        db.commit()
    return RedirectResponse("/executors", 302)


# ===================== ANALITIKA =====================
@app.get("/analytics", response_class=HTMLResponse)
def analytics_page(request: Request, db: Session = Depends(get_db)):
    user = current_user(request, db)
    if not user:
        return RedirectResponse("/login", 302)
    if not has_perm(user, "view_analytics"):
        return RedirectResponse("/dashboard", 302)
    base = db.query(models.Request)
    total = base.count()
    status_stats = {s: base.filter(models.Request.status == s).count() for s in Status}
    done = status_stats[Status.done]
    sla = round(done / total * 100) if total else 0
    overdue = base.filter(models.Request.deadline.isnot(None),
                         models.Request.deadline < datetime.utcnow(),
                         models.Request.status.notin_([Status.done, Status.rejected])).count()
    dep_stats = (db.query(models.Department.name, models.Department.color,
                          func.count(models.Request.id))
                 .outerjoin(models.Request).group_by(models.Department.id).all())
    max_count = max([row[2] for row in dep_stats], default=1) or 1
    # ijrochilar samaradorligi
    execs = db.query(models.User).filter(models.User.is_active == True).all()
    exec_stats = []
    for e in execs:
        rq = db.query(models.Request).filter(models.Request.assignees.any(models.User.id == e.id))
        t = rq.count()
        d = rq.filter(models.Request.status == Status.done).count()
        if t:
            exec_stats.append({"name": e.full_name, "email": e.email, "total": t,
                               "done": d, "pct": round(d / t * 100)})
    return templates.TemplateResponse(request, "analytics.html", {
        "request": request, "user": user, "active": "analytics",
        "total": total, "sla": sla, "overdue": overdue,
        "status_stats": {s.value: status_stats[s] for s in Status},
        "dep_stats": dep_stats, "max_count": max_count, "exec_stats": exec_stats,
    })


# ===================== BO'LIMLAR =====================
@app.get("/departments", response_class=HTMLResponse)
def departments_page(request: Request, error: Optional[str] = None,
                     db: Session = Depends(get_db)):
    user = current_user(request, db)
    if not user:
        return RedirectResponse("/login", 302)
    return templates.TemplateResponse(request, "departments.html",
        {"request": request, "user": user, "active": "departments",
         "departments": db.query(models.Department).all(), "error": error})


@app.post("/departments/create")
def departments_create(request: Request, name: str = Form(...), icon: str = Form("🗂️"),
                       color: str = Form("#2563eb"), db: Session = Depends(get_db)):
    user = current_user(request, db)
    if not user:
        return RedirectResponse("/login", 302)
    if db.query(models.Department).count() >= 10:
        return RedirectResponse("/departments?error=limit", 302)
    db.add(models.Department(name=name.strip(), icon=(icon.strip() or "🗂️"), color=color))
    db.commit()
    return RedirectResponse("/departments", 302)


@app.post("/departments/{dep_id}/delete")
def departments_delete(dep_id: int, request: Request, db: Session = Depends(get_db)):
    user = current_user(request, db)
    if not user:
        return RedirectResponse("/login", 302)
    d = db.get(models.Department, dep_id)
    if d:
        db.query(models.User).filter(models.User.department_id == dep_id)\
            .update({models.User.department_id: None}, synchronize_session=False)
        for r in db.query(models.Request).filter(models.Request.department_id == dep_id).all():
            db.query(models.Notification).filter(
                models.Notification.link == f"/requests/{r.id}").delete(synchronize_session=False)
            db.delete(r)
        db.delete(d); db.commit()
    return RedirectResponse("/departments", 302)


# ===================== ADMIN PANEL =====================
@app.get("/admin", response_class=HTMLResponse)
def admin_page(request: Request, db: Session = Depends(get_db)):
    user = current_user(request, db)
    if not user:
        return RedirectResponse("/login", 302)
    if user.role != Role.admin:
        return RedirectResponse("/dashboard", 302)
    from sqlalchemy import func as _func
    req_counts = dict(db.query(models.Request.created_by, _func.count(models.Request.id))
                      .group_by(models.Request.created_by).all())
    # har filialning login (client) emaili — tahrirlashda ko'rsatish uchun
    branch_logins = dict(db.query(models.User.user_branch_id, models.User.email)
                         .filter(models.User.role == Role.client,
                                 models.User.user_branch_id.isnot(None)).all())
    return templates.TemplateResponse(request, "admin.html", {
        "request": request, "user": user, "active": "admin",
        "users": db.query(models.User).order_by(models.User.full_name).all(),
        "departments": db.query(models.Department).all(),
        "branches": db.query(models.Branch).order_by(models.Branch.name).all(),
        "req_counts": req_counts, "branch_logins": branch_logins,
    })


@app.post("/admin/users/create")
def admin_user_create(request: Request, full_name: str = Form(...), email: str = Form(...),
                      password: str = Form("12345678"), role: str = Form("executor"),
                      department_id: Optional[int] = Form(None), phone: str = Form(""),
                      telegram_chat_id: str = Form(""), kpp_branch_ids: List[int] = Form([]),
                      perms: List[str] = Form([]),
                      db: Session = Depends(get_db)):
    user = current_user(request, db)
    if not user or user.role != Role.admin:
        return RedirectResponse("/login", 302)
    email = email.lower().strip()
    if role == "kpp":
        department_id = None
    if not db.query(models.User).filter(models.User.email == email).first():
        nu = models.User(full_name=full_name.strip(), email=email,
                         hashed_password=auth.hash_password(password or "12345678"),
                         role=Role(role), department_id=department_id,
                         phone=phone.strip(), telegram_chat_id=telegram_chat_id.strip(),
                         is_active=True)
        if role == "kpp" and kpp_branch_ids:
            nu.visible_branches = db.query(models.Branch).filter(
                models.Branch.id.in_(kpp_branch_ids)).all()
        # ruxsatlar belgilangan bo'lsa — aniq yozamiz; aks holda null (rol bo'yicha)
        if perms:
            nu.perms = json.dumps({k: (k in perms) for k in PERMISSION_KEYS})
        db.add(nu)
        db.commit()
    return RedirectResponse("/admin", 302)


@app.post("/admin/users/{uid}/delete")
def admin_user_delete(uid: int, request: Request, db: Session = Depends(get_db)):
    user = current_user(request, db)
    if not user or user.role != Role.admin:
        return RedirectResponse("/login", 302)
    p = db.get(models.User, uid)
    if p and p.id != user.id:
        # bu odam ijrochi bo'lgan заявкаларни bo'shatamiz
        db.query(models.Request).filter(models.Request.assigned_to == uid)\
            .update({models.Request.assigned_to: None}, synchronize_session=False)
        # bu odam yaratgan заявкаларni birma-bir o'chiramiz (cascade: izoh/tarix/fayl)
        reqs = db.query(models.Request).filter(models.Request.created_by == uid).all()
        for r in reqs:
            for att in r.attachments:
                try:
                    fp = att.file_path.lstrip("/")
                    if os.path.exists(fp):
                        os.remove(fp)
                except Exception:
                    pass
            db.query(models.Notification).filter(
                models.Notification.link == f"/requests/{r.id}").delete(synchronize_session=False)
            db.delete(r)
        # bu odamning izohlari va bildirishnomalarini o'chiramiz
        db.query(models.Comment).filter(models.Comment.user_id == uid)\
            .delete(synchronize_session=False)
        db.query(models.Notification).filter(models.Notification.user_id == uid)\
            .delete(synchronize_session=False)
        db.delete(p)
        db.commit()
    return RedirectResponse("/admin", 302)


@app.post("/admin/users/{uid}/edit")
def admin_user_edit(uid: int, request: Request, full_name: str = Form(...),
                    email: str = Form(""), role: str = Form("executor"),
                    department_id: Optional[int] = Form(None),
                    phone: str = Form(""), telegram_chat_id: str = Form(""),
                    password: str = Form(""), kpp_branch_ids: List[int] = Form([]),
                    perms: List[str] = Form([]), db: Session = Depends(get_db)):
    user = current_user(request, db)
    if not user or user.role != Role.admin:
        return RedirectResponse("/login", 302)
    p = db.get(models.User, uid)
    if p:
        p.full_name = full_name.strip()
        # email — boshqa odamda band bo'lmasa o'zgartiramiz
        new_email = email.lower().strip()
        if new_email and new_email != p.email:
            taken = db.query(models.User).filter(
                models.User.email == new_email, models.User.id != uid).first()
            if not taken:
                p.email = new_email
        p.role = Role(role)
        p.department_id = None if role == "kpp" else department_id
        p.phone = phone.strip()
        p.telegram_chat_id = telegram_chat_id.strip()
        # КПП filiallari
        if role == "kpp":
            p.visible_branches = db.query(models.Branch).filter(
                models.Branch.id.in_(kpp_branch_ids)).all() if kpp_branch_ids else []
        else:
            p.visible_branches = []
        # parol — faqat kiritilgan bo'lsa o'zgartiramiz
        if password.strip():
            p.hashed_password = auth.hash_password(password.strip())
        # ruxsatlar (admin belgilaydi) — barcha kalitlar aniq yoziladi
        p.perms = json.dumps({k: (k in perms) for k in PERMISSION_KEYS})
        db.commit()
    return RedirectResponse("/admin", 302)


@app.post("/admin/requests/clear")
def admin_clear_requests(request: Request, db: Session = Depends(get_db)):
    """Barcha zayavkalar, kommentlar, fayllar va tegishli bildirishnomalarni o'chiradi."""
    user = current_user(request, db)
    if not user or user.role != Role.admin:
        return RedirectResponse("/login", 302)
    # biriktirilgan fayllarni diskdan o'chiramiz
    for a in db.query(models.Attachment).all():
        fp = (a.file_path or "").lstrip("/")
        try:
            if fp and os.path.exists(fp):
                os.remove(fp)
        except Exception:
            pass
    db.query(models.Attachment).delete(synchronize_session=False)
    db.query(models.Comment).delete(synchronize_session=False)
    db.query(models.StatusHistory).delete(synchronize_session=False)
    db.query(models.Notification).filter(
        models.Notification.link.like("/requests/%")).delete(synchronize_session=False)
    # stop-list tarixini (yechilgan yozuvlar) ham tozalaymiz
    db.query(models.StopEntry).filter(
        models.StopEntry.resolved == True).delete(synchronize_session=False)
    try:
        db.execute(models.request_assignees.delete())
    except Exception:
        pass
    db.query(models.Request).delete(synchronize_session=False)
    db.commit()
    print(">>> [MAXWAY] Barcha zayavkalar tozalandi", flush=True)
    return RedirectResponse("/admin?cleared=1", 302)
def admin_cat_create(request: Request, name: str = Form(...), icon: str = Form("🗂️"),
                     color: str = Form("#2563eb"), subcategories: str = Form(""),
                     db: Session = Depends(get_db)):
    user = current_user(request, db)
    if not user or user.role != Role.admin:
        return RedirectResponse("/login", 302)
    if db.query(models.Department).count() < 10:
        d = models.Department(name=name.strip(), icon=(icon.strip() or "🗂️"), color=color)
        db.add(d); db.flush()
        for line in subcategories.splitlines():
            nm = line.strip()
            if nm:
                db.add(models.Subcategory(name=nm, department_id=d.id))
        db.commit()
    return RedirectResponse("/admin", 302)


@app.post("/admin/categories/{dep_id}/delete")
def admin_cat_delete(dep_id: int, request: Request, db: Session = Depends(get_db)):
    user = current_user(request, db)
    if not user or user.role != Role.admin:
        return RedirectResponse("/login", 302)
    d = db.get(models.Department, dep_id)
    if d:
        db.query(models.User).filter(models.User.department_id == dep_id)\
            .update({models.User.department_id: None}, synchronize_session=False)
        for r in db.query(models.Request).filter(models.Request.department_id == dep_id).all():
            for att in r.attachments:
                try:
                    fp = att.file_path.lstrip("/")
                    if os.path.exists(fp):
                        os.remove(fp)
                except Exception:
                    pass
            db.query(models.Notification).filter(
                models.Notification.link == f"/requests/{r.id}").delete(synchronize_session=False)
            db.delete(r)
        db.delete(d); db.commit()
    return RedirectResponse("/admin", 302)


@app.post("/admin/categories/{dep_id}/edit")
def admin_cat_edit(dep_id: int, request: Request, name: str = Form(...),
                   icon: str = Form("🗂️"), color: str = Form("#2563eb"),
                   subcategories: str = Form(""), db: Session = Depends(get_db)):
    user = current_user(request, db)
    if not user or user.role != Role.admin:
        return RedirectResponse("/login", 302)
    d = db.get(models.Department, dep_id)
    if d:
        d.name = name.strip()
        d.icon = icon.strip() or "🗂️"
        d.color = color
        # podkategoriyalarni qayta yozamiz (avval заявкалардаги ishorani bo'shatamiz)
        old_ids = [s.id for s in db.query(models.Subcategory).filter(
            models.Subcategory.department_id == dep_id).all()]
        if old_ids:
            db.query(models.Request).filter(
                models.Request.subcategory_id.in_(old_ids)).update(
                {models.Request.subcategory_id: None}, synchronize_session=False)
        db.query(models.Subcategory).filter(
            models.Subcategory.department_id == dep_id).delete()
        for line in subcategories.splitlines():
            nm = line.strip()
            if nm:
                db.add(models.Subcategory(name=nm, department_id=dep_id))
        db.commit()
    return RedirectResponse("/admin", 302)


@app.post("/admin/branches/create")
def admin_branch_create(request: Request, name: str = Form(...), location: str = Form(""),
                        phone: str = Form(""), director_name: str = Form(""),
                        login_email: str = Form(""), password: str = Form(""),
                        tg_chat_ids: str = Form(""),
                        db: Session = Depends(get_db)):
    user = current_user(request, db)
    if not user or user.role != Role.admin:
        return RedirectResponse("/login", 302)
    b = models.Branch(name=name.strip(), location=location.strip(),
                      phone=phone.strip(), director_name=director_name.strip(),
                      tg_chat_ids=tg_chat_ids.strip())
    db.add(b); db.flush()
    # filial uchun login (klient) yaratish
    login_email = login_email.lower().strip()
    if login_email and not db.query(models.User).filter(models.User.email == login_email).first():
        db.add(models.User(full_name=name.strip(), email=login_email,
                           hashed_password=auth.hash_password(password or "12345678"),
                           role=Role.client, user_branch_id=b.id, is_active=True))
    db.commit()
    return RedirectResponse("/admin", 302)


@app.post("/admin/branches/{bid}/delete")
def admin_branch_delete(bid: int, request: Request, db: Session = Depends(get_db)):
    user = current_user(request, db)
    if not user or user.role != Role.admin:
        return RedirectResponse("/login", 302)
    b = db.get(models.Branch, bid)
    if b:
        db.query(models.Request).filter(models.Request.branch_id == bid)\
            .update({models.Request.branch_id: None})
        db.delete(b); db.commit()
    return RedirectResponse("/admin", 302)


@app.post("/admin/branches/{bid}/edit")
def admin_branch_edit(bid: int, request: Request, name: str = Form(...),
                      location: str = Form(""), phone: str = Form(""),
                      director_name: str = Form(""), login_email: str = Form(""),
                      password: str = Form(""), tg_chat_ids: str = Form(""),
                      db: Session = Depends(get_db)):
    user = current_user(request, db)
    if not user or user.role != Role.admin:
        return RedirectResponse("/login", 302)
    b = db.get(models.Branch, bid)
    if b:
        b.name = name.strip()
        b.location = location.strip()
        b.phone = phone.strip()
        b.director_name = director_name.strip()
        b.tg_chat_ids = tg_chat_ids.strip()
        # filial logini (client akkaunt)
        login_email = login_email.lower().strip()
        client = db.query(models.User).filter(
            models.User.user_branch_id == bid, models.User.role == Role.client).first()
        if login_email:
            taken = db.query(models.User).filter(
                models.User.email == login_email,
                models.User.user_branch_id != bid).first()
            if client:
                # mavjud login: email (band bo'lmasa) va parol (kiritilgan bo'lsa) yangilanadi
                if not taken and login_email != client.email:
                    client.email = login_email
                if password.strip():
                    client.hashed_password = auth.hash_password(password.strip())
            elif not taken:
                # login yo'q edi — yangi client akkaunt yaratamiz
                db.add(models.User(full_name=b.name, email=login_email,
                                   hashed_password=auth.hash_password(password.strip() or "12345678"),
                                   role=Role.client, user_branch_id=bid, is_active=True))
        elif client and password.strip():
            # email o'zgarmasa ham — faqat parol yangilash
            client.hashed_password = auth.hash_password(password.strip())
        db.commit()
    return RedirectResponse("/admin", 302)


# ===================== JSON API =====================
@app.get("/api/departments")
def api_departments(db: Session = Depends(get_db)):
    return [{"id": d.id, "name": d.name, "icon": d.icon, "color": d.color,
             "count": len(d.requests)} for d in db.query(models.Department).all()]


@app.get("/api/requests")
def api_requests(department_id: Optional[int] = None, db: Session = Depends(get_db)):
    q = db.query(models.Request)
    if department_id:
        q = q.filter(models.Request.department_id == department_id)
    return [{"id": r.id, "title": r.title, "status": r.status.value,
             "priority": r.priority.value,
             "department": r.department.name if r.department else None,
             "assignee": ", ".join(a.full_name for a in r.assignees) if r.assignees else None}
            for r in q.order_by(models.Request.created_at.desc()).all()]


# ===================== STOP-LIST =====================
def is_supply(user):
    """Foydalanuvchi Снабжение bo'limidami?"""
    return user.department is not None and "набжен" in (user.department.name or "").lower()


# ===================== RUXSATLAR (admin har bir userga belgilaydi) =====================
# (kalit, ko'rinadigan nomi, bo'lim)
PERMISSION_DEFS = [
    ("create_request", "Создавать заявки", "Заявки"),
    ("assign",         "Назначать исполнителей", "Заявки"),
    ("add_stop",       "Добавлять в стоп-лист", "Стоп-лист"),
    ("edit_stop",      "Редактировать запись стоп-листа", "Стоп-лист"),
    ("resolve_stop",   "Убирать из стоп-листа", "Стоп-лист"),
    ("comment_stop",   "Комментировать стоп-лист (снабжение)", "Стоп-лист"),
    ("confirm_stop",   "Подтверждать причину стопа (снабжение)", "Стоп-лист"),
    ("export_stop",    "Выгружать стоп-лист в Excel", "Стоп-лист"),
    ("manage_menu",    "Управлять меню стоп-листа", "Стоп-лист"),
    ("view_analytics", "Видеть аналитику", "Аналитика"),
]
PERMISSION_KEYS = [k for k, _, _ in PERMISSION_DEFS]


def _role_default(user, key):
    """Ruxsat maxsus belgilanmagan bo'lsa — rol bo'yicha standart xatti-harakat."""
    r = user.role
    if key == "create_request":
        return r in (Role.admin, Role.manager, Role.client)
    if key == "assign":
        return r in (Role.admin, Role.manager)
    if key == "add_stop":
        return r == Role.client
    if key == "edit_stop":
        # filial — o'z yozuvini, Снабжение — o'z maydonlarini tahrirlaydi
        return r == Role.client or is_supply(user)
    if key == "resolve_stop":
        return r == Role.client or is_supply(user)
    if key == "comment_stop":
        return is_supply(user)
    if key == "confirm_stop":
        # stop sababini tasdiqlash — faqat Снабжение (va admin, has_perm'da)
        return is_supply(user)
    if key == "export_stop":
        return r in (Role.client, Role.viewer) or is_supply(user)
    if key == "manage_menu":
        return r == Role.manager and is_supply(user)
    if key == "view_analytics":
        return r in (Role.admin, Role.viewer)
    return False


def has_perm(user, key):
    """Admin — hamma narsaga ruxsatli. Aks holda: maxsus ruxsat bo'lsa o'sha, bo'lmasa rol standarti."""
    if user is None:
        return False
    if user.role == Role.admin:
        return True
    try:
        perms = json.loads(user.perms) if user.perms else None
    except Exception:
        perms = None
    if isinstance(perms, dict) and key in perms:
        return bool(perms[key])
    return _role_default(user, key)


# shablonlarда `can(user, 'create_request')` sifatida ishlatiladi
templates.env.globals["can"] = has_perm
templates.env.globals["PERMISSION_DEFS"] = PERMISSION_DEFS


def can_see_stoplist(user):
    # Снабжение, filial direktorlari, viewer, admin
    return user.role in (Role.client, Role.viewer, Role.admin) or is_supply(user)


def can_manage_menu(user):
    """Menyu boshqaruvi — ruxsatga qarab (default: Снабжение menejeri)."""
    return has_perm(user, "manage_menu")


# ---------- Stop-list: ruxsatlar (maydonlar darajasida) ----------
def can_edit_stop_branch_fields(user, e) -> bool:
    """Filial maydonlari (причина, комментарий филиала) — o'z filiali yoki admin."""
    if user.role == Role.admin:
        return True
    if not has_perm(user, "edit_stop"):
        return False
    if user.role == Role.client:
        return e.branch_id == user.user_branch_id
    return False


def can_edit_stop_supply_comment(user, e) -> bool:
    """Снабжение izohi."""
    return has_perm(user, "comment_stop")


def can_confirm_stop(user, e) -> bool:
    """Подтверждение причины стопа отделом снабжения (ДА/НЕТ)."""
    return has_perm(user, "confirm_stop")


def can_view_stop(user, e) -> bool:
    """Yozuvni ko'rish: filial faqat o'zinikini, qolganlar — ruxsatga qarab."""
    if not can_see_stoplist(user):
        return False
    if user.role == Role.client:
        return e.branch_id == user.user_branch_id
    return True


# ---------- Stop-list: filtr / saralash / sahifalash ----------
def _stop_filters(user, branch_id="", menu_item_id="", reason="", confirmed="",
                  date_from="", date_to="", month="") -> dict:
    """Query paramlarni tozalaydi. Noto'g'ri qiymat — e'tiborsiz qoldiriladi (filtr yo'q)."""
    f = {"branch_id": None, "menu_item_id": None, "reason": "", "confirmed": "",
         "date_from": "", "date_to": "", "month": ""}
    # filial: klient doim o'z filiali bilan cheklanadi (backendda majburiy)
    if user.role == Role.client:
        f["branch_id"] = user.user_branch_id
    elif str(branch_id).isdigit():
        f["branch_id"] = int(branch_id)
    if str(menu_item_id).isdigit():
        f["menu_item_id"] = int(menu_item_id)
    if reason in REASON_LABELS:
        f["reason"] = reason
    if confirmed in ("yes", "no"):
        f["confirmed"] = confirmed
    for key, val in (("date_from", date_from), ("date_to", date_to)):
        val = (val or "").strip()
        if val:
            try:
                datetime.strptime(val, "%Y-%m-%d")
                f[key] = val
            except ValueError:
                pass
    month = (month or "").strip()
    if month:
        try:
            datetime.strptime(month, "%Y-%m")
            f["month"] = month
        except ValueError:
            pass
    return f


def _stop_query(db: Session, user, f: dict, resolved: bool):
    """Filtrlangan StopEntry query. branch/menu_item darhol yuklanadi (N+1 yo'q)."""
    from sqlalchemy.orm import contains_eager
    q = (db.query(models.StopEntry)
         .outerjoin(models.Branch, models.StopEntry.branch_id == models.Branch.id)
         .outerjoin(models.MenuItem, models.StopEntry.menu_item_id == models.MenuItem.id)
         .options(contains_eager(models.StopEntry.branch),
                  contains_eager(models.StopEntry.menu_item))
         .filter(models.StopEntry.resolved == resolved))
    # klient filialga biriktirilmagan bo'lsa — hech nima ko'rmaydi (ma'lumot sizib chiqmasin)
    if user.role == Role.client and not f["branch_id"]:
        return q.filter(sqltext("1 = 0"))
    if f["branch_id"]:
        q = q.filter(models.StopEntry.branch_id == f["branch_id"])
    if f["menu_item_id"]:
        q = q.filter(models.StopEntry.menu_item_id == f["menu_item_id"])
    if f["reason"]:
        q = q.filter(models.StopEntry.reason == f["reason"])
    if f["confirmed"]:
        q = q.filter(models.StopEntry.supply_confirmed == (f["confirmed"] == "yes"))
    if f["month"]:
        m0 = datetime.strptime(f["month"], "%Y-%m")
        m1 = m0.replace(year=m0.year + 1, month=1) if m0.month == 12 \
            else m0.replace(month=m0.month + 1)
        q = q.filter(models.StopEntry.created_at >= m0, models.StopEntry.created_at < m1)
    if f["date_from"]:
        q = q.filter(models.StopEntry.created_at >= datetime.strptime(f["date_from"], "%Y-%m-%d"))
    if f["date_to"]:
        q = q.filter(models.StopEntry.created_at
                     < datetime.strptime(f["date_to"], "%Y-%m-%d") + timedelta(days=1))
    return q


def _stop_sorted(q, sort_by: str, sort_order: str, resolved: bool):
    """Saralash. Faqat oq ro'yxatdagi ustunlar — SQL injection imkonsiz."""
    if sort_by not in STOP_SORT_FIELDS:
        sort_by = "resolved_at" if resolved else "created_at"
    order = "asc" if sort_order == "asc" else "desc"
    col = STOP_SORT_FIELDS[sort_by]()
    expr = col.asc() if order == "asc" else col.desc()
    if sort_by == "resolved_at":
        expr = expr.nullslast()
    # barqaror tartib uchun ikkilamchi kalit
    return q.order_by(expr, models.StopEntry.id.desc()), sort_by, order


def _stop_page(q, page: str, page_size: str):
    """Sahifalash. Qaytadi: (yozuvlar, meta-dict)."""
    try:
        size = int(page_size)
    except (TypeError, ValueError):
        size = STOP_PAGE_SIZE_DEFAULT
    if size not in STOP_PAGE_SIZES:
        size = STOP_PAGE_SIZE_DEFAULT
    total = q.order_by(None).count()
    pages = max(1, -(-total // size))
    try:
        cur = int(page)
    except (TypeError, ValueError):
        cur = 1
    cur = min(max(cur, 1), pages)
    items = q.offset((cur - 1) * size).limit(size).all()
    return items, {"total": total, "page": cur, "page_size": size, "pages": pages}


def _name_key(s: str):
    """Alifbo tartibi uchun kalit. SQLite'ning ORDER BY'i baytlar bo'yicha saralaydi:
    kichik harflar va «Ё» ro'yxat oxiriga tushib qoladi. Bu kalit buni to'g'rilaydi."""
    return (s or "").casefold().replace("ё", "е")


def sorted_by_name(items):
    """Nomi bo'yicha to'g'ri alifbo tartibi (lotin/kirill, katta-kichik harf aralash)."""
    return sorted(items, key=lambda x: _name_key(x.name))


def _validate_stop(db: Session, branch_id, menu_item_id, reason,
                   comment="", supply_comment=""):
    """Backend validatsiyasi. Qaytadi: (xatolar ro'yxati, tozalangan qiymatlar)."""
    errors = []
    if not branch_id:
        errors.append("Филиал обязателен")
    elif not db.get(models.Branch, branch_id):
        errors.append("Филиал не найден")
    if not menu_item_id:
        errors.append("Блюдо обязательно")
    elif not db.get(models.MenuItem, menu_item_id):
        errors.append("Блюдо не найдено")
    if not reason:
        errors.append("Причина обязательна")
    elif reason not in REASON_LABELS:
        errors.append("Причина не найдена в справочнике")
    comment = (comment or "").strip()
    supply_comment = (supply_comment or "").strip()
    if len(comment) > STOP_COMMENT_MAX:
        errors.append(f"Комментарий филиала — максимум {STOP_COMMENT_MAX} символов")
    if len(supply_comment) > STOP_COMMENT_MAX:
        errors.append(f"Комментарий снабжения — максимум {STOP_COMMENT_MAX} символов")
    return errors, {"comment": comment, "supply_comment": supply_comment}


def _parse_bool(v) -> bool:
    """ДА/НЕТ qiymatini boolean'ga aylantiradi (form ham, JSON ham)."""
    if isinstance(v, bool):
        return v
    return str(v).strip().lower() in ("1", "true", "yes", "on", "да")


def _stop_dict(e) -> dict:
    """Bitta yozuvning JSON ko'rinishi (API va detal sahifa uchun)."""
    return {
        "id": e.id,
        "created_at": e.created_at.strftime("%Y-%m-%d %H:%M") if e.created_at else None,
        "branch_id": e.branch_id,
        "branch": e.branch.name if e.branch else None,
        "product_id": e.menu_item_id,
        "product": e.menu_item.name if e.menu_item else None,
        "reason": e.reason,
        "reason_label": REASON_LABELS.get(e.reason, e.reason),
        "branch_comment": e.comment or "",
        "supply_confirmed": bool(e.supply_confirmed),
        "supply_comment": e.supply_comment or "",
        "created_by": e.created_by,
        "created_by_name": display_name(e.creator) if e.creator else None,
        "updated_at": e.updated_at.strftime("%Y-%m-%d %H:%M") if e.updated_at else None,
        "updated_by": e.updated_by,
        "updated_by_name": display_name(e.updater) if e.updater else None,
        "confirmed_at": e.confirmed_at.strftime("%Y-%m-%d %H:%M") if e.confirmed_at else None,
        "confirmed_by_name": display_name(e.confirmer) if e.confirmer else None,
        "resolved": bool(e.resolved),
        "resolved_at": e.resolved_at.strftime("%Y-%m-%d %H:%M") if e.resolved_at else None,
    }


def _touch_stop(e, user):
    """Audit: har qanday o'zgarishда kim va qachon o'zgartirgani yoziladi."""
    e.updated_by = user.id
    e.updated_at = models.tashkent_now()


def _stoplist_context(request: Request, db: Session, user, resolved: bool,
                      branch_id, menu_item_id, reason, confirmed,
                      date_from, date_to, month, sort_by, sort_order, page, page_size):
    """/stoplist va /stoplist/history uchun umumiy kontekst (filtr+sort+pagination)."""
    is_client = user.role == Role.client
    f = _stop_filters(user, branch_id, menu_item_id, reason, confirmed,
                      date_from, date_to, month)
    q = _stop_query(db, user, f, resolved)
    q, sort_by, sort_order = _stop_sorted(q, sort_by, sort_order, resolved)
    entries, pg = _stop_page(q, page, page_size)
    # filtr uchun spravochniklar
    branches = [] if is_client else sorted_by_name(db.query(models.Branch).all())
    dishes = sorted_by_name(db.query(models.MenuItem).all())
    # joriy filtrni saqlab qoluvchi query-string (sort/pagination/export havolalari uchun)
    keep = {"branch_id": f["branch_id"] if not is_client else "",
            "menu_item_id": f["menu_item_id"], "reason": f["reason"],
            "confirmed": f["confirmed"], "date_from": f["date_from"],
            "date_to": f["date_to"], "month": f["month"], "page_size": pg["page_size"]}
    qs = urllib.parse.urlencode({k: v for k, v in keep.items() if v not in ("", None)})
    # klientning filiali majburiy filtr — uni "foydalanuvchi qo'ygan filtr" deb hisoblamaymiz
    has_filters = bool(f["menu_item_id"] or f["reason"] or f["confirmed"]
                       or f["date_from"] or f["date_to"] or f["month"]
                       or (f["branch_id"] and not is_client))
    return {
        "request": request, "user": user, "entries": entries, "is_client": is_client,
        "branches": branches, "dishes": dishes, "f": f, "pg": pg,
        "has_filters": has_filters,
        "sort_by": sort_by, "sort_order": sort_order, "qs": qs,
        "cur_month": models.tashkent_now().strftime("%Y-%m"),
        "can_add": has_perm(user, "add_stop"),
        "can_resolve": has_perm(user, "resolve_stop"),
        "can_comment": has_perm(user, "comment_stop"),
        "can_confirm": has_perm(user, "confirm_stop"),
        "can_edit": has_perm(user, "edit_stop") or user.role == Role.admin,
        "can_export": has_perm(user, "export_stop"),
    }


@app.get("/stoplist", response_class=HTMLResponse)
def stoplist_page(request: Request, sync: str = "", err: str = "", ok: str = "",
                  branch_id: str = "", menu_item_id: str = "", reason: str = "",
                  confirmed: str = "", date_from: str = "", date_to: str = "",
                  month: str = "", sort_by: str = "created_at", sort_order: str = "desc",
                  page: str = "1", page_size: str = "", db: Session = Depends(get_db)):
    user = current_user(request, db)
    if not user:
        return RedirectResponse("/login", 302)
    if not can_see_stoplist(user):
        return RedirectResponse("/dashboard", 302)
    ctx = _stoplist_context(request, db, user, False, branch_id, menu_item_id, reason,
                            confirmed, date_from, date_to, month,
                            sort_by, sort_order, page, page_size)
    # qo'shish formasi uchun faqat faol menyu (alifbo tartibida)
    ctx["menu_items"] = sorted_by_name(db.query(models.MenuItem).filter(
        models.MenuItem.is_active == True).all())
    ctx.update({"active": "stoplist", "sync_msg": sync, "err_msg": err, "ok_msg": ok})
    return templates.TemplateResponse(request, "stoplist.html", ctx)


@app.get("/stoplist/history", response_class=HTMLResponse)
def stoplist_history_page(request: Request, err: str = "", ok: str = "",
                          branch_id: str = "", menu_item_id: str = "", reason: str = "",
                          confirmed: str = "", date_from: str = "", date_to: str = "",
                          month: str = "", sort_by: str = "resolved_at",
                          sort_order: str = "desc", page: str = "1", page_size: str = "",
                          db: Session = Depends(get_db)):
    user = current_user(request, db)
    if not user:
        return RedirectResponse("/login", 302)
    if not can_see_stoplist(user):
        return RedirectResponse("/dashboard", 302)
    ctx = _stoplist_context(request, db, user, True, branch_id, menu_item_id, reason,
                            confirmed, date_from, date_to, month,
                            sort_by, sort_order, page, page_size)
    ctx.update({"active": "stophistory", "err_msg": err, "ok_msg": ok,
                "history": ctx["entries"]})
    return templates.TemplateResponse(request, "stoplist_history.html", ctx)


@app.get("/stoplist/menu", response_class=HTMLResponse)
def stoplist_menu_page(request: Request, sync: str = "", db: Session = Depends(get_db)):
    user = current_user(request, db)
    if not user:
        return RedirectResponse("/login", 302)
    if not can_manage_menu(user):
        return RedirectResponse("/dashboard", 302)
    menu_items = sorted_by_name(db.query(models.MenuItem).all())
    return templates.TemplateResponse(request, "stoplist_menu.html", {
        "request": request, "user": user, "active": "stopmenu",
        "menu_items": menu_items, "sync_msg": sync,
    })


def _create_stop_entries(db: Session, user, branch_id, item_ids, reason,
                         comment, supply_comment="", supply_confirmed=False):
    """Bir yoki bir nechta taomni stopga qo'yadi.
    created_at / created_by — faqat backend belgilaydi (frontendga ishonilmaydi).
    Qaytadi: (yaratilgan yozuvlar, xatolar, allaqachon stopda bo'lganlar)."""
    if not item_ids:
        return [], ["Блюдо обязательно"], []
    created, skipped, errors = [], [], []
    for mid in item_ids:
        errs, clean = _validate_stop(db, branch_id, mid, reason, comment, supply_comment)
        if errs:
            return [], errs, []
        mi = db.get(models.MenuItem, mid)
        exists = db.query(models.StopEntry).filter(
            models.StopEntry.branch_id == branch_id,
            models.StopEntry.menu_item_id == mid,
            models.StopEntry.resolved == False).first()
        if exists:
            skipped.append(mi.name)
            continue
        e = models.StopEntry(
            branch_id=branch_id, menu_item_id=mid, reason=reason,
            comment=clean["comment"], supply_comment=clean["supply_comment"],
            supply_confirmed=bool(supply_confirmed),
            created_by=user.id, created_at=models.tashkent_now())
        if supply_confirmed:
            e.confirmed_by = user.id
            e.confirmed_at = models.tashkent_now()
        db.add(e)
        created.append(e)
    if created:
        db.commit()
        notify_stop_added(db, created, user)
    return created, errors, skipped


def stop_notify_targets(db: Session, branch_id: int, actor=None) -> List[str]:
    """Stop qo'shilganda kimga telegram ketishini aniqlaydi.

    Barcha filiallar bo'yicha: Снабжение, Просмотр (viewer), Админ.
    КПП — admin biriktirgan filiallar bo'yicha (biriktirilmagan bo'lsa — barchasi).
    Faqat o'z filiali bo'yicha: filial logini (client) va filialga yozilgan
    qo'shimcha telegram ID lar (Branch.tg_chat_ids).
    Xabarni qo'shgan odamning o'ziga yuborilmaydi."""
    ids, seen = [], set()

    def _add(cid):
        cid = (cid or "").strip()
        if cid and cid not in seen:
            seen.add(cid)
            ids.append(cid)

    users = db.query(models.User).filter(
        models.User.is_active == True,
        models.User.telegram_chat_id.isnot(None),
        models.User.telegram_chat_id != "").all()
    for u in users:
        if actor and u.id == actor.id:
            continue
        if u.role == Role.client:
            if u.user_branch_id and u.user_branch_id == branch_id:
                _add(u.telegram_chat_id)
        elif u.role == Role.kpp:
            vis = {b.id for b in u.visible_branches}
            if not vis or branch_id in vis:
                _add(u.telegram_chat_id)
        elif u.role in (Role.viewer, Role.admin) or is_supply(u):
            _add(u.telegram_chat_id)
    # filialning qo'shimcha xodimlari (login emas — faqat telegram ID)
    b = db.get(models.Branch, branch_id) if branch_id else None
    if b and b.tg_chat_ids:
        for part in b.tg_chat_ids.replace(";", ",").replace("\n", ",").replace(" ", ",").split(","):
            _add(part)
    # umumiy kanal/guruh (sozlangan bo'lsa)
    _add(get_stop_channel())
    return ids


def notify_stop_added(db: Session, created, actor):
    """Filialdan stopga taom qo'shilganda telegram xabari (bitta umumiy xabar)."""
    if not created:
        return
    first = created[0]
    branch = first.branch or db.get(models.Branch, first.branch_id)
    dishes = [(e.menu_item.name if e.menu_item else "—") for e in created]
    shown = dishes[:15]
    more = len(dishes) - len(shown)
    lines = [f"🛑 <b>Новый стоп — MAXWAY</b>", "",
             f"🏢 Филиал: <b>{branch.name if branch else '—'}</b>",
             f"🏷 Причина: {REASON_LABELS.get(first.reason, first.reason)}",
             f"🍽 Блюда ({len(dishes)}):"]
    lines += [f" • {n}" for n in shown]
    if more > 0:
        lines.append(f" • …и ещё {more}")
    if first.comment:
        lines.append(f"💬 Комментарий филиала: {first.comment}")
    lines.append(f"👤 Добавил: {display_name(actor)}")
    lines.append(f"🕑 {first.created_at.strftime('%d.%m.%Y %H:%M')}"
                 if first.created_at else "")
    text = "\n".join(l for l in lines if l != "")
    link = f"/stoplist/{first.id}" if len(created) == 1 else "/stoplist"
    _send_async(stop_notify_targets(db, first.branch_id, actor),
                text, button_url=f"{get_app_url()}{link}",
                token=get_stop_bot_token())


def _human_duration(delta) -> str:
    """Muddatni ruscha qisqa yozadi: «2 д 5 ч», «3 ч 40 мин», «12 мин»."""
    mins = int(delta.total_seconds() // 60)
    if mins < 1:
        return "меньше минуты"
    d, rem = divmod(mins, 1440)
    h, m = divmod(rem, 60)
    parts = []
    if d:
        parts.append(f"{d} д")
    if h:
        parts.append(f"{h} ч")
    if m and not d:            # kunlar bo'lsa daqiqa ortiqcha
        parts.append(f"{m} мин")
    return " ".join(parts) or f"{mins} мин"


def notify_stop_resolved(db: Session, e, actor):
    """Taom stopdan olinganda telegram xabari — qo'shilgandagi bilan bir xil manzillarga."""
    if not e:
        return
    branch = e.branch or db.get(models.Branch, e.branch_id)
    lines = ["✅ <b>Снят со стопа — MAXWAY</b>", "",
             f"🏢 Филиал: <b>{branch.name if branch else '—'}</b>",
             f"🍽 Блюдо: <b>{e.menu_item.name if e.menu_item else '—'}</b>",
             f"🏷 Причина была: {REASON_LABELS.get(e.reason, e.reason)}"]
    if e.created_at and e.resolved_at:
        lines.append(f"⏱ На стопе был: {_human_duration(e.resolved_at - e.created_at)}")
    if e.supply_comment:
        lines.append(f"💬 Комментарий снабжения: {e.supply_comment}")
    lines.append(f"👤 Снял: {display_name(actor)}")
    if e.resolved_at:
        lines.append(f"🕑 {e.resolved_at.strftime('%d.%m.%Y %H:%M')}")
    _send_async(stop_notify_targets(db, e.branch_id, actor),
                "\n".join(lines), button_url=f"{get_app_url()}/stoplist/{e.id}",
                token=get_stop_bot_token())


@app.post("/stoplist/add")
def stoplist_add(request: Request, menu_item_id: List[int] = Form([]),
                 menu_name: List[str] = Form([]), branch_id: str = Form(""),
                 reason: str = Form(""), comment: str = Form(""),
                 db: Session = Depends(get_db)):
    user = current_user(request, db)
    if not user:
        return RedirectResponse("/login", 302)
    if not has_perm(user, "add_stop"):
        return RedirectResponse("/stoplist?err=" + urllib.parse.quote(
            "Нет прав на добавление в стоп-лист"), 302)
    # filial: klient — doim o'ziniki; admin/менеджер — formadan
    if user.role == Role.client:
        b_id = user.user_branch_id
    else:
        b_id = int(branch_id) if str(branch_id).isdigit() else None
    # taomlar: id bo'yicha (yangi forma) yoki nomi bo'yicha (eski forma)
    ids = list(menu_item_id)
    for nm in menu_name:
        mi = db.query(models.MenuItem).filter(
            models.MenuItem.name == (nm or "").strip(),
            models.MenuItem.is_active == True).first()
        if mi and mi.id not in ids:
            ids.append(mi.id)
    created, errors, skipped = _create_stop_entries(db, user, b_id, ids, reason, comment)
    if errors:
        # xato bo'lsa — qo'shish sahifasiga qaytamiz, foydalanuvchi tuzatsin
        return RedirectResponse("/stoplist/new?err=" + urllib.parse.quote("; ".join(errors)), 302)
    msg = f"Добавлено записей: {len(created)}"
    if skipped:
        msg += f" · уже в стоп-листе: {', '.join(skipped[:5])}"
    return RedirectResponse("/stoplist?ok=" + urllib.parse.quote(msg), 302)


def _apply_stop_update(db: Session, user, e, data: dict):
    """Yozuvni tahrirlash — har bir maydon o'z ruxsatiga tekshiriladi.
    Qaytadi: (o'zgardimi, xato matni yoki None, HTTP kod)."""
    changed = False
    # --- filial maydonlari: причина, комментарий филиала ---
    wants_branch = any(k in data for k in ("reason", "branch_comment"))
    if wants_branch:
        if not can_edit_stop_branch_fields(user, e):
            return False, "Нет прав на редактирование полей филиала", 403
        reason = data.get("reason", e.reason)
        comment = data.get("branch_comment", e.comment)
        errs, clean = _validate_stop(db, e.branch_id, e.menu_item_id, reason, comment)
        if errs:
            return False, "; ".join(errs), 400
        if e.reason != reason:
            e.reason = reason; changed = True
        if (e.comment or "") != clean["comment"]:
            e.comment = clean["comment"]; changed = True
    # --- снабжение izohi ---
    if "supply_comment" in data:
        if not can_edit_stop_supply_comment(user, e):
            return False, "Нет прав на комментарий снабжения", 403
        val = (data["supply_comment"] or "").strip()
        if len(val) > STOP_COMMENT_MAX:
            return False, f"Комментарий снабжения — максимум {STOP_COMMENT_MAX} символов", 400
        if (e.supply_comment or "") != val:
            e.supply_comment = val; changed = True
    # --- подтверждение причины стопа (ДА/НЕТ) ---
    if "supply_confirmed" in data:
        if not can_confirm_stop(user, e):
            return False, "Нет прав на подтверждение причины стопа", 403
        val = _parse_bool(data["supply_confirmed"])
        if bool(e.supply_confirmed) != val:
            e.supply_confirmed = val
            e.confirmed_by = user.id if val else None
            e.confirmed_at = models.tashkent_now() if val else None
            changed = True
    if changed:
        _touch_stop(e, user)
        db.commit()
    return changed, None, 200


@app.post("/stoplist/{sid}/edit")
def stoplist_edit(sid: int, request: Request,
                  reason: str = Form(None), branch_comment: str = Form(None),
                  supply_comment: str = Form(None), supply_confirmed: str = Form(None),
                  fields: str = Form(""), db: Session = Depends(get_db)):
    """Yozuvni tahrirlash. `fields` — qaysi maydonlar yuborilgani (vergul bilan)."""
    user = current_user(request, db)
    if not user:
        return RedirectResponse("/login", 302)
    e = db.get(models.StopEntry, sid)
    if not e:
        raise HTTPException(status_code=404, detail="Запись не найдена")
    if not can_view_stop(user, e):
        raise HTTPException(status_code=403, detail="Нет доступа к этой записи")
    sent = {n.strip() for n in fields.split(",") if n.strip()}
    data = {}
    for name, val in (("reason", reason), ("branch_comment", branch_comment),
                      ("supply_comment", supply_comment),
                      ("supply_confirmed", supply_confirmed)):
        # checkbox yuborilmasa None keladi — shuning uchun `fields` ro'yxatiga tayanamiz
        if name in sent or (val is not None and not sent):
            data[name] = val if val is not None else ""
    _, err, code = _apply_stop_update(db, user, e, data)
    if err:
        raise HTTPException(status_code=code, detail=err)
    return RedirectResponse(f"/stoplist/{sid}?ok=" + urllib.parse.quote("Сохранено"), 302)


@app.post("/stoplist/{sid}/comment")
def stoplist_comment(sid: int, request: Request, comment: str = Form(""),
                     db: Session = Depends(get_db)):
    """Снабжение xodimi o'z izohini (supply_comment) qo'shadi/tahrirlaydi."""
    user = current_user(request, db)
    if not user:
        return RedirectResponse("/login", 302)
    e = db.get(models.StopEntry, sid)
    if not e:
        raise HTTPException(status_code=404, detail="Запись не найдена")
    _, err, code = _apply_stop_update(db, user, e, {"supply_comment": comment})
    if err:
        raise HTTPException(status_code=code, detail=err)
    ref = request.headers.get("referer") or "/stoplist"
    dest = "/stoplist/history" if "history" in ref else "/stoplist"
    return RedirectResponse(dest, 302)


@app.post("/stoplist/{sid}/confirm")
def stoplist_confirm(sid: int, request: Request, supply_confirmed: str = Form("0"),
                     supply_comment: str = Form(None), db: Session = Depends(get_db)):
    """Подтверждение причины стопа отделом снабжения (ДА/НЕТ)."""
    user = current_user(request, db)
    if not user:
        return RedirectResponse("/login", 302)
    e = db.get(models.StopEntry, sid)
    if not e:
        raise HTTPException(status_code=404, detail="Запись не найдена")
    data = {"supply_confirmed": supply_confirmed}
    if supply_comment is not None:
        data["supply_comment"] = supply_comment
    _, err, code = _apply_stop_update(db, user, e, data)
    if err:
        raise HTTPException(status_code=code, detail=err)
    ref = request.headers.get("referer") or "/stoplist"
    return RedirectResponse(ref, 302)


@app.post("/stoplist/{sid}/resolve")
def stoplist_resolve(sid: int, request: Request, db: Session = Depends(get_db)):
    user = current_user(request, db)
    if not user:
        return RedirectResponse("/login", 302)
    e = db.get(models.StopEntry, sid)
    if e:
        # o'z filiali (client) yoki resolve ruxsati borlar
        ok = (user.role == Role.client and e.branch_id == user.user_branch_id) \
            or has_perm(user, "resolve_stop")
        if not ok:
            raise HTTPException(status_code=403, detail="Нет прав убирать из стоп-листа")
        e.resolved = True
        e.resolved_at = models.tashkent_now()
        _touch_stop(e, user)
        db.commit()
        notify_stop_resolved(db, e, user)
    return RedirectResponse("/stoplist", 302)


@app.post("/stoplist/menu/add")
def stoplist_menu_add(request: Request, name: str = Form(...), db: Session = Depends(get_db)):
    user = current_user(request, db)
    if not user or not can_manage_menu(user):
        return RedirectResponse("/login", 302)
    for line in name.splitlines():
        nm = line.strip()
        if nm and not db.query(models.MenuItem).filter(models.MenuItem.name == nm).first():
            db.add(models.MenuItem(name=nm))
    db.commit()
    return RedirectResponse("/stoplist/menu", 302)


@app.post("/stoplist/menu/{mid}/delete")
def stoplist_menu_delete(mid: int, request: Request, db: Session = Depends(get_db)):
    user = current_user(request, db)
    if not user or not can_manage_menu(user):
        return RedirectResponse("/login", 302)
    m = db.get(models.MenuItem, mid)
    if m:
        db.query(models.StopEntry).filter(models.StopEntry.menu_item_id == mid).delete()
        db.delete(m); db.commit()
    return RedirectResponse("/stoplist/menu", 302)


# ===================== MENYU YUKLASH (Excel/PDF) + STOP-LIST EXPORT =====================
def _parse_menu_file(filename, data):
    """Excel (.xlsx) yoki PDF fayldan taom nomlarini ajratib oladi. Nomlar ro'yxatini qaytaradi."""
    names = []
    fn = (filename or "").lower()
    if fn.endswith((".xlsx", ".xlsm")):
        import openpyxl, io
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if isinstance(cell, str) and cell.strip() and len(cell.strip()) >= 2:
                        names.append(cell.strip())
        wb.close()
    elif fn.endswith(".pdf"):
        from pypdf import PdfReader
        import io
        reader = PdfReader(io.BytesIO(data))
        for page in reader.pages:
            for line in (page.extract_text() or "").splitlines():
                s = line.strip()
                if len(s) >= 2:
                    names.append(s)
    else:
        raise ValueError("Faqat .xlsx yoki .pdf")
    # takrorlarni olib tashlaymiz (tartibni saqlab)
    seen, out = set(), []
    for n in names:
        if n not in seen:
            seen.add(n); out.append(n)
    return out


@app.post("/stoplist/menu/upload")
def stoplist_menu_upload(request: Request, replace: str = Form(""),
                         file: UploadFile = File(...), db: Session = Depends(get_db)):
    user = current_user(request, db)
    if not user or not can_manage_menu(user):
        return RedirectResponse("/login", 302)
    try:
        data = file.file.read()
        names = _parse_menu_file(file.filename, data)
    except Exception as e:
        return RedirectResponse(f"/stoplist/menu?sync={urllib.parse.quote('Ошибка файла: ' + str(e)[:120])}", 302)
    if replace == "1":
        # eski menyuni o'chiramiz (stop-listlar bilan)
        db.query(models.StopEntry).delete(synchronize_session=False)
        db.query(models.MenuItem).delete(synchronize_session=False)
    added = 0
    for nm in names:
        if not db.query(models.MenuItem).filter(models.MenuItem.name == nm).first():
            db.add(models.MenuItem(name=nm, is_active=True)); added += 1
    db.commit()
    msg = f"Меню загружено: {added} новых из {len(names)} (файл: {file.filename})"
    return RedirectResponse(f"/stoplist/menu?sync={urllib.parse.quote(msg)}", 302)


@app.post("/stoplist/menu/clear")
def stoplist_menu_clear(request: Request, db: Session = Depends(get_db)):
    user = current_user(request, db)
    if not user or not can_manage_menu(user):
        return RedirectResponse("/login", 302)
    db.query(models.StopEntry).delete(synchronize_session=False)
    db.query(models.MenuItem).delete(synchronize_session=False)
    db.commit()
    return RedirectResponse("/stoplist/menu?sync=" + urllib.parse.quote("Меню очищено"), 302)


def _xlsx_title(name, used):
    bad = '[]:*?/\\'
    t = "".join(c for c in (name or "—") if c not in bad)[:28] or "Лист"
    base = t; i = 2
    while t in used:
        t = f"{base[:25]} {i}"; i += 1
    used.add(t); return t


@app.get("/stoplist/export")
def stoplist_export(request: Request, mode: str = "active",
                    branch_id: str = "", menu_item_id: str = "", reason: str = "",
                    confirmed: str = "", date_from: str = "", date_to: str = "",
                    month: str = "", sort_by: str = "", sort_order: str = "desc",
                    db: Session = Depends(get_db)):
    """Excel eksport — ekrandagi filtrlar va saralash aynan qo'llanadi."""
    user = current_user(request, db)
    if not user or not can_see_stoplist(user):
        return RedirectResponse("/login", 302)
    if not has_perm(user, "export_stop"):
        raise HTTPException(status_code=403, detail="Нет прав на выгрузку в Excel")
    import openpyxl, io
    from openpyxl.styles import Font, PatternFill, Alignment
    from sqlalchemy.orm import joinedload
    from fastapi.responses import StreamingResponse
    resolved = (mode == "history")
    f = _stop_filters(user, branch_id, menu_item_id, reason, confirmed,
                      date_from, date_to, month)
    q = _stop_query(db, user, f, resolved).options(
        joinedload(models.StopEntry.creator))
    if not sort_by:
        sort_by = "resolved_at" if resolved else "created_at"
    q, _sb, _so = _stop_sorted(q, sort_by, sort_order, resolved)
    entries = q.all()

    headers = ["Добавлено", "Филиал", "Блюдо", "Причина", "Комментарий Филиала",
               "Подтверждение причины стопа отделом снабжения", "Комментарий Снабжения"]
    widths = [18, 24, 34, 28, 30, 22, 30]
    if resolved:
        headers.append("Убрано"); widths.append(18)

    def row(e):
        r = [e.created_at.strftime("%d.%m.%Y %H:%M") if e.created_at else "",
             e.branch.name if e.branch else "—",
             e.menu_item.name if e.menu_item else "—",
             REASON_LABELS.get(e.reason, e.reason),
             e.comment or "",
             CONFIRM_LABELS[bool(e.supply_confirmed)],
             e.supply_comment or ""]
        if resolved:
            r.append(e.resolved_at.strftime("%d.%m.%Y %H:%M") if e.resolved_at else "")
        return r

    head_fill = PatternFill("solid", fgColor="1E293B")
    head_font = Font(color="FFFFFF", bold=True)
    head_align = Alignment(vertical="center", wrap_text=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    def fill(title, items):
        ws = wb.create_sheet(title)
        ws.append(headers)
        for c in ws[1]:
            c.fill = head_fill; c.font = head_font; c.alignment = head_align
        for e in items:
            ws.append(row(e))
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws.freeze_panes = "A2"

    fill("Все филиалы", entries)
    # har bir filial uchun alohida varaq (klient bo'lmasa)
    if user.role != Role.client:
        by_branch = {}
        for e in entries:
            bn = e.branch.name if e.branch else "—"
            by_branch.setdefault(bn, []).append(e)
        used = {"Все филиалы"}
        for bn in sorted(by_branch):
            fill(_xlsx_title(bn, used), by_branch[bn])

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    tag = "product_stops_history" if resolved else "product_stops"
    fname = f"{tag}_{models.tashkent_now().strftime('%Y-%m-%d')}.xlsx"
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})


@app.get("/stoplist/new", response_class=HTMLResponse)
def stoplist_new_page(request: Request, err: str = "", db: Session = Depends(get_db)):
    """Stopga qo'shish — alohida to'liq sahifa."""
    user = current_user(request, db)
    if not user:
        return RedirectResponse("/login", 302)
    if not has_perm(user, "add_stop"):
        return RedirectResponse("/stoplist?err=" + urllib.parse.quote(
            "Нет прав на добавление в стоп-лист"), 302)
    is_client = user.role == Role.client
    menu_items = sorted_by_name(db.query(models.MenuItem).filter(
        models.MenuItem.is_active == True).all())
    # allaqachon stopda turgan taomlar — qayta tanlanmasin
    br_id = user.user_branch_id if is_client else None
    on_stop = set()
    if br_id:
        on_stop = {r[0] for r in db.query(models.StopEntry.menu_item_id).filter(
            models.StopEntry.branch_id == br_id,
            models.StopEntry.resolved == False).all()}
    return templates.TemplateResponse(request, "stoplist_new.html", {
        "request": request, "user": user, "active": "stoplist",
        "is_client": is_client, "menu_items": menu_items, "on_stop": on_stop,
        "branches": [] if is_client else sorted_by_name(db.query(models.Branch).all()),
        "my_branch": db.get(models.Branch, br_id) if br_id else None,
        "err_msg": err,
    })


@app.get("/stoplist/{sid}", response_class=HTMLResponse)
def stoplist_detail(sid: int, request: Request, ok: str = "", err: str = "",
                    db: Session = Depends(get_db)):
    """Yozuv haqida to'liq ma'lumot (просмотр) + ruxsat bo'lsa tahrirlash formasi."""
    user = current_user(request, db)
    if not user:
        return RedirectResponse("/login", 302)
    from sqlalchemy.orm import joinedload
    e = (db.query(models.StopEntry)
         .options(joinedload(models.StopEntry.branch),
                  joinedload(models.StopEntry.menu_item),
                  joinedload(models.StopEntry.creator),
                  joinedload(models.StopEntry.updater),
                  joinedload(models.StopEntry.confirmer))
         .filter(models.StopEntry.id == sid).first())
    if not e:
        raise HTTPException(status_code=404, detail="Запись не найдена")
    if not can_view_stop(user, e):
        raise HTTPException(status_code=403, detail="Нет доступа к этой записи")
    return templates.TemplateResponse(request, "stoplist_detail.html", {
        "request": request, "user": user,
        "active": "stophistory" if e.resolved else "stoplist",
        "e": e, "ok_msg": ok, "err_msg": err,
        "can_edit_branch": can_edit_stop_branch_fields(user, e),
        "can_edit_supply": can_edit_stop_supply_comment(user, e),
        "can_confirm": can_confirm_stop(user, e),
        "can_resolve": (user.role == Role.client and e.branch_id == user.user_branch_id)
                       or has_perm(user, "resolve_stop"),
    })


# ===================== STOP-LIST JSON API =====================
@app.get("/api/product-stops")
def api_product_stops(request: Request, mode: str = "active",
                      branch_id: str = "", menu_item_id: str = "", reason: str = "",
                      confirmed: str = "", date_from: str = "", date_to: str = "",
                      month: str = "", sort_by: str = "created_at",
                      sort_order: str = "desc", page: str = "1", page_size: str = "",
                      db: Session = Depends(get_db)):
    """Stop-list ro'yxati: filtr + saralash + sahifalash."""
    user = current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Требуется авторизация")
    if not can_see_stoplist(user):
        raise HTTPException(status_code=403, detail="Нет доступа к стоп-листу")
    resolved = (mode == "history")
    f = _stop_filters(user, branch_id, menu_item_id, reason, confirmed,
                      date_from, date_to, month)
    q = _stop_query(db, user, f, resolved)
    q, sb, so = _stop_sorted(q, sort_by, sort_order, resolved)
    items, pg = _stop_page(q, page, page_size)
    return JSONResponse({"items": [_stop_dict(e) for e in items],
                         "sort_by": sb, "sort_order": so, **pg})


@app.get("/api/product-stops/{sid}")
def api_product_stop_detail(sid: int, request: Request, db: Session = Depends(get_db)):
    user = current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Требуется авторизация")
    e = db.get(models.StopEntry, sid)
    if not e:
        raise HTTPException(status_code=404, detail="Запись не найдена")
    if not can_view_stop(user, e):
        raise HTTPException(status_code=403, detail="Нет доступа к этой записи")
    return JSONResponse(_stop_dict(e))


@app.post("/api/product-stops")
async def api_product_stop_create(request: Request, db: Session = Depends(get_db)):
    """Yangi stop yozuvi. created_at/created_by — faqat backend belgilaydi."""
    user = current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Требуется авторизация")
    if not has_perm(user, "add_stop"):
        raise HTTPException(status_code=403, detail="Нет прав на добавление в стоп-лист")
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Ожидается JSON")
    if not isinstance(body, dict):
        raise HTTPException(status_code=400, detail="Ожидается JSON-объект")
    if user.role == Role.client:
        b_id = user.user_branch_id
    else:
        b_id = body.get("branch_id")
        b_id = int(b_id) if str(b_id).isdigit() else None
    raw_ids = body.get("product_ids") or ([body["product_id"]] if body.get("product_id") else [])
    ids = [int(i) for i in raw_ids if str(i).isdigit()]
    confirmed = _parse_bool(body.get("supply_confirmed", False))
    if confirmed and not has_perm(user, "confirm_stop"):
        raise HTTPException(status_code=403, detail="Нет прав на подтверждение причины стопа")
    created, errors, skipped = _create_stop_entries(
        db, user, b_id, ids, body.get("reason", ""), body.get("branch_comment", ""),
        body.get("supply_comment", ""), confirmed)
    if errors:
        raise HTTPException(status_code=400, detail="; ".join(errors))
    return JSONResponse({"created": [_stop_dict(e) for e in created],
                         "skipped": skipped}, status_code=201)


@app.patch("/api/product-stops/{sid}")
async def api_product_stop_update(sid: int, request: Request, db: Session = Depends(get_db)):
    """Yozuvni tahrirlash. Har bir maydon alohida ruxsatga tekshiriladi."""
    user = current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Требуется авторизация")
    e = db.get(models.StopEntry, sid)
    if not e:
        raise HTTPException(status_code=404, detail="Запись не найдена")
    if not can_view_stop(user, e):
        raise HTTPException(status_code=403, detail="Нет доступа к этой записи")
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Ожидается JSON")
    if not isinstance(body, dict):
        raise HTTPException(status_code=400, detail="Ожидается JSON-объект")
    allowed = ("reason", "branch_comment", "supply_comment", "supply_confirmed")
    data = {k: v for k, v in body.items() if k in allowed}
    if not data:
        raise HTTPException(status_code=400, detail="Нет полей для изменения")
    _, err, code = _apply_stop_update(db, user, e, data)
    if err:
        raise HTTPException(status_code=code, detail=err)
    db.refresh(e)
    return JSONResponse(_stop_dict(e))


# PUT — PATCH bilan bir xil semantika (mavjud maydonlar yangilanadi)
app.add_api_route("/api/product-stops/{sid}", api_product_stop_update, methods=["PUT"])


@app.get("/api/stop-reasons")
def api_stop_reasons(request: Request, db: Session = Depends(get_db)):
    """Sabablar spravochnigi (dropdown uchun)."""
    user = current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Требуется авторизация")
    return JSONResponse([{"id": k, "name": v} for k, v in REASON_LABELS.items()])


@app.get("/api/branches")
def api_branches(request: Request, db: Session = Depends(get_db)):
    """Filiallar spravochnigi. Klient — faqat o'z filialini ko'radi."""
    user = current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Требуется авторизация")
    q = db.query(models.Branch)
    if user.role == Role.client and user.user_branch_id:
        q = q.filter(models.Branch.id == user.user_branch_id)
    return JSONResponse([{"id": b.id, "name": b.name, "location": b.location or ""}
                         for b in sorted_by_name(q.all())])


@app.get("/api/products")
def api_products(request: Request, active_only: str = "1", db: Session = Depends(get_db)):
    """Blyudolar (menyu) spravochnigi."""
    user = current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Требуется авторизация")
    q = db.query(models.MenuItem)
    if active_only == "1":
        q = q.filter(models.MenuItem.is_active == True)
    return JSONResponse([{"id": m.id, "name": m.name}
                         for m in sorted_by_name(q.all())])
